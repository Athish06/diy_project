"""
rule_matcher.py — all FastAPI routes, safety analysis, DB access, cache,
extraction pipeline, and evaluation logic in one place.

Endpoints:
  GET  /api/health
  GET  /api/analyze          (SSE) — full analysis pipeline
  GET  /api/rules
  GET  /api/filter_options
  GET  /api/rules_by_document
  GET  /api/extraction_runs
  POST /api/extract_rules
  POST /api/run_evaluation/{run_id}
  POST /api/scans
  GET  /api/scans
  GET  /api/scans/{scan_id}
  WS   /ws/extract           — WebSocket extraction with progress
"""

# ---------------------------------------------------------------------------
# stdlib / third-party imports
# ---------------------------------------------------------------------------
import asyncio
import base64
import csv
import io
import json
import logging
import math
import os
import random
import re
import shutil
import subprocess
import tempfile
import time
from urllib.parse import parse_qs, urlparse
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional, List

import httpx
import psycopg2
import psycopg2.extras
from fastapi import (
    APIRouter, HTTPException, Query, UploadFile, File, Form,
    WebSocket, WebSocketDisconnect, Request,
)
from sse_starlette.sse import EventSourceResponse

from transcript import fetch_transcript, fetch_metadata
from steps_extract import extract_steps_stream
from embeddings import EmbeddingService

logger = logging.getLogger("diy.rule_matcher")

# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------

def get_api_key() -> str:
    return os.getenv("GROQ_API_KEY", "")


def get_model() -> str:
    return os.getenv("MODEL", "qwen/qwen3-32b")


def get_database_url() -> str:
    return os.getenv("DATABASE_URL") or os.getenv("SUPABASE_URL", "")


# ---------------------------------------------------------------------------
# In-memory cache (24 h TTL, max 200 entries)
# ---------------------------------------------------------------------------

CACHE_TTL_S = 86_400
CACHE_MAX_SIZE = 200


class _CacheEntry:
    __slots__ = ("data", "created_at")

    def __init__(self, data: str):
        self.data = data
        self.created_at = time.monotonic()


class AnalysisCache:
    def __init__(self):
        self._entries: dict[str, _CacheEntry] = {}

    def get(self, video_id: str) -> Optional[str]:
        entry = self._entries.get(video_id)
        if entry is None:
            return None
        if time.monotonic() - entry.created_at > CACHE_TTL_S:
            del self._entries[video_id]
            return None
        return entry.data

    def set(self, video_id: str, data: str):
        self._entries[video_id] = _CacheEntry(data)
        self._cleanup()

    def _cleanup(self):
        now = time.monotonic()
        expired = [k for k, v in self._entries.items() if now - v.created_at > CACHE_TTL_S]
        for k in expired:
            del self._entries[k]
        if len(self._entries) > CACHE_MAX_SIZE:
            sorted_keys = sorted(self._entries, key=lambda k: self._entries[k].created_at)
            for k in sorted_keys[: len(self._entries) - CACHE_MAX_SIZE]:
                del self._entries[k]


# ---------------------------------------------------------------------------
# DB connection
# ---------------------------------------------------------------------------

def get_db_connection():
    raw_url = os.getenv("DATABASE_URL") or os.getenv("SUPABASE_URL", "")
    if not raw_url:
        raise RuntimeError("DATABASE_URL or SUPABASE_URL not set")
    url = raw_url
    return psycopg2.connect(url)


# ---------------------------------------------------------------------------
# DB queries
# ---------------------------------------------------------------------------

def fetch_rules_from_db(
    category: str | None = None,
    severity: int | None = None,
    document: str | None = None,
    search: str | None = None,
    page: int = 1,
    per_page: int = 50,
) -> dict:
    conn = get_db_connection()
    try:
        conditions = []
        params: list[Any] = []
        run_ids_for_document: list[int] | None = None

        if category:
            conditions.append("%s = ANY(sr.categories)")
            params.append(category)
        if severity is not None:
            conditions.append("sr.validated_severity = %s")
            params.append(severity)
        if document:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM extraction_runs WHERE json_source_file = %s",
                    (document,),
                )
                run_ids_for_document = [int(r[0]) for r in cur.fetchall()]

            # Enforce filename -> run_id mapping first, then fetch from safety_rules by run_id.
            if run_ids_for_document:
                conditions.append("sr.run_id = ANY(%s)")
                params.append(run_ids_for_document)
            else:
                # Backward-compatible fallback for legacy rows that may not have run linkage.
                conditions.append("COALESCE(er.json_source_file, sr.source_document) = %s")
                params.append(document)
        if search:
            conditions.append("(sr.actionable_rule ILIKE %s OR sr.original_text ILIKE %s)")
            params.extend([f"%{search}%", f"%{search}%"])

        where_clause = (" WHERE " + " AND ".join(conditions)) if conditions else ""

        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT COUNT(*)
                FROM safety_rules sr
                LEFT JOIN extraction_runs er ON er.id = sr.run_id
                {where_clause}
                """,
                params,
            )
            total = cur.fetchone()[0]

        offset = (page - 1) * per_page
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                f"""
              SELECT sr.id, sr.rule_id, sr.original_text, sr.actionable_rule, sr.materials,
                  sr.suggested_severity, sr.validated_severity, sr.categories,
                      COALESCE(er.json_source_file, sr.source_document) AS source_document,
                  sr.page_number, sr.section_heading, sr.run_id, sr.created_at
                  FROM safety_rules sr
                  LEFT JOIN extraction_runs er ON er.id = sr.run_id
                  {where_clause}
                  ORDER BY sr.validated_severity DESC, sr.id
                LIMIT %s OFFSET %s
                """,
                params + [per_page, offset],
            )
            rows = cur.fetchall()

        rules = []
        for row in rows:
            r = dict(row)
            if r.get("created_at"):
                r["created_at"] = r["created_at"].isoformat()
            if r.get("rule_id"):
                r["rule_id"] = str(r["rule_id"])
            rules.append(r)

        return {"rules": rules, "total": total, "page": page, "per_page": per_page}
    finally:
        conn.close()


def fetch_filter_options_from_db() -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT DISTINCT unnest(categories) AS cat FROM safety_rules ORDER BY cat"
            )
            categories = [row[0] for row in cur.fetchall()]

            cur.execute(
                "SELECT DISTINCT validated_severity FROM safety_rules "
                "WHERE validated_severity IS NOT NULL ORDER BY validated_severity DESC"
            )
            severities = [row[0] for row in cur.fetchall()]

            cur.execute(
                """
                SELECT DISTINCT COALESCE(er.json_source_file, sr.source_document) AS source_document
                FROM safety_rules sr
                LEFT JOIN extraction_runs er ON er.id = sr.run_id
                ORDER BY source_document
                """
            )
            documents = [row[0] for row in cur.fetchall()]

        return {"categories": categories, "severities": severities, "documents": documents}
    finally:
        conn.close()


def fetch_rules_by_document() -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT
                    COALESCE(er.json_source_file, sr.source_document) AS name,
                    COUNT(*) AS rule_count,
                    ARRAY_AGG(DISTINCT unnest_cat) AS categories,
                    ROUND(AVG(sr.validated_severity)::numeric, 1) AS avg_severity,
                    MAX(sr.created_at) AS last_updated
                FROM safety_rules sr
                LEFT JOIN extraction_runs er ON er.id = sr.run_id
                CROSS JOIN LATERAL unnest(sr.categories) AS unnest_cat
                GROUP BY COALESCE(er.json_source_file, sr.source_document)
                ORDER BY COALESCE(er.json_source_file, sr.source_document)
            """)
            rows = cur.fetchall()

        documents = []
        for row in rows:
            r = dict(row)
            if r.get("last_updated"):
                r["last_updated"] = r["last_updated"].isoformat()
            if r.get("avg_severity"):
                r["avg_severity"] = float(r["avg_severity"])
            documents.append(r)

        return {"documents": documents}
    finally:
        conn.close()


def fetch_extraction_runs() -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id, run_timestamp, model_used, total_pages, rule_count,
                       document_count, source_documents, json_source_file,
                       file_url, evaluation_results, created_at
                FROM extraction_runs
                ORDER BY id DESC
            """)
            rows = cur.fetchall()

        runs = []
        for row in rows:
            r = dict(row)
            if r.get("run_timestamp"):
                r["run_timestamp"] = r["run_timestamp"].isoformat()
            if r.get("created_at"):
                r["created_at"] = r["created_at"].isoformat()
            if r.get("evaluation_results") and isinstance(r["evaluation_results"], str):
                r["evaluation_results"] = json.loads(r["evaluation_results"])
            runs.append(r)

        return {"runs": runs}
    finally:
        conn.close()


def fetch_rules_by_run(run_id: int, page: int = 1, per_page: int = 50) -> dict:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM safety_rules WHERE run_id = %s", (run_id,))
            total = cur.fetchone()[0]

        offset = (page - 1) * per_page
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
              SELECT sr.id, sr.rule_id, sr.original_text, sr.actionable_rule, sr.materials,
                  sr.suggested_severity, sr.validated_severity, sr.categories,
                      COALESCE(er.json_source_file, sr.source_document) AS source_document,
                  sr.page_number, sr.section_heading, sr.run_id, sr.created_at
                  FROM safety_rules sr
                  LEFT JOIN extraction_runs er ON er.id = sr.run_id
                  WHERE sr.run_id = %s
                  ORDER BY sr.validated_severity DESC, sr.id
                LIMIT %s OFFSET %s
                """,
                (run_id, per_page, offset),
            )
            rows = cur.fetchall()

        rules = []
        for row in rows:
            r = dict(row)
            if r.get("created_at"):
                r["created_at"] = r["created_at"].isoformat()
            if r.get("rule_id"):
                r["rule_id"] = str(r["rule_id"])
            rules.append(r)

        return {"rules": rules, "total": total, "page": page, "per_page": per_page}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# DB migrations
# ---------------------------------------------------------------------------

MIGRATIONS = [
    """
    DO $$
    BEGIN
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'safety_rules' AND column_name = 'run_id'
        ) THEN
            ALTER TABLE safety_rules ADD COLUMN run_id INTEGER REFERENCES extraction_runs(id);
        END IF;
    END $$;
    """,
    "UPDATE safety_rules SET run_id = 1 WHERE run_id IS NULL;",
    "CREATE INDEX IF NOT EXISTS idx_rules_run_id ON safety_rules (run_id);",
    """
    DO $$
    BEGIN
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'extraction_runs' AND column_name = 'evaluation_results'
        ) THEN
            ALTER TABLE extraction_runs ADD COLUMN evaluation_results JSONB;
        END IF;
    END $$;
    """,
    """
    DO $$
    BEGIN
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'extraction_runs' AND column_name = 'file_url'
        ) THEN
            ALTER TABLE extraction_runs ADD COLUMN file_url TEXT;
        END IF;
    END $$;
    """,
    """
    DO $$
    BEGIN
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'completed_scans' AND column_name = 'model_reports'
        ) THEN
            ALTER TABLE completed_scans ADD COLUMN model_reports JSONB;
        END IF;
    END $$;
    """,
    """
    DO $$
    BEGIN
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'completed_scans' AND column_name = 'comparison_data'
        ) THEN
            ALTER TABLE completed_scans ADD COLUMN comparison_data JSONB;
        END IF;
    END $$;
    """,
    """
    CREATE TABLE IF NOT EXISTS evaluation_results (
        id                          SERIAL PRIMARY KEY,
        run_id                      INTEGER NOT NULL REFERENCES extraction_runs(id),
        file_name                   TEXT NOT NULL,
        total_rules                 INTEGER DEFAULT 0,
        text_presence_passed        INTEGER DEFAULT 0,
        text_presence_total         INTEGER DEFAULT 0,
        page_accuracy_passed        INTEGER DEFAULT 0,
        page_accuracy_total         INTEGER DEFAULT 0,
        heading_accuracy_passed     INTEGER DEFAULT 0,
        heading_accuracy_total      INTEGER DEFAULT 0,
        category_validity_passed    INTEGER DEFAULT 0,
        category_validity_total     INTEGER DEFAULT 0,
        severity_consistency_passed INTEGER DEFAULT 0,
        severity_consistency_total  INTEGER DEFAULT 0,
        hallucination_rate          REAL,
        correctness_score           REAL,
        overall_accuracy            REAL,
        failed_rules                JSONB,
        created_at                  TIMESTAMPTZ DEFAULT now()
    );
    """,
    "CREATE INDEX IF NOT EXISTS idx_eval_run_id ON evaluation_results (run_id);",
    """
    CREATE TABLE IF NOT EXISTS system_eval (
        id                      SERIAL PRIMARY KEY,
        evaluated_at            TIMESTAMPTZ DEFAULT now(),
        model_key               TEXT DEFAULT 'qwen',
        sample_size             INTEGER DEFAULT 0,
        evaluated_scans         INTEGER DEFAULT 0,
        total_steps             INTEGER DEFAULT 0,
        total_precautions       INTEGER DEFAULT 0,
        supported_precautions   INTEGER DEFAULT 0,
        true_positive           INTEGER DEFAULT 0,
        true_negative           INTEGER DEFAULT 0,
        false_positive          INTEGER DEFAULT 0,
        false_negative          INTEGER DEFAULT 0,
        accuracy                REAL,
        precision               REAL,
        recall                  REAL,
        f1_score                REAL,
        mean_reciprocal_rank    REAL,
        faithfulness_score      REAL,
        spearman_correlation    REAL,
        details_json            JSONB
    );
    """,
    """
    DO $$
    BEGIN
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'youtube_urls'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN youtube_urls JSONB;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'selected_urls_count'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN selected_urls_count INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'total_urls_in_pool'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN total_urls_in_pool INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_total_steps'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_total_steps INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_total_precautions'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_total_precautions INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_supported_precautions'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_supported_precautions INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_true_positive'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_true_positive INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_true_negative'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_true_negative INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_false_positive'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_false_positive INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_false_negative'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_false_negative INTEGER DEFAULT 0;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_accuracy'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_accuracy REAL;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_precision'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_precision REAL;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_recall'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_recall REAL;
        END IF;
        IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'system_eval' AND column_name = 'cum_f1_score'
        ) THEN
            ALTER TABLE system_eval ADD COLUMN cum_f1_score REAL;
        END IF;
    END $$;
    """,
    """
    CREATE TABLE IF NOT EXISTS youtube_urls (
        id              SERIAL PRIMARY KEY,
        url             TEXT NOT NULL UNIQUE,
        video_id        TEXT,
        source_type     TEXT DEFAULT 'manual',
        source_file     TEXT,
        created_at      TIMESTAMPTZ DEFAULT now(),
        last_used_at    TIMESTAMPTZ
    );
    """,
    "CREATE INDEX IF NOT EXISTS idx_youtube_urls_video_id ON youtube_urls (video_id);",
    """
    CREATE TABLE IF NOT EXISTS system_eval_video_results (
        id                      SERIAL PRIMARY KEY,
        eval_id                 INTEGER NOT NULL REFERENCES system_eval(id) ON DELETE CASCADE,
        video_id                TEXT,
        video_url               TEXT,
        scan_id                 INTEGER,
        steps_evaluated         INTEGER DEFAULT 0,
        total_precautions       INTEGER DEFAULT 0,
        supported_precautions   INTEGER DEFAULT 0,
        true_positive           INTEGER DEFAULT 0,
        true_negative           INTEGER DEFAULT 0,
        false_positive          INTEGER DEFAULT 0,
        false_negative          INTEGER DEFAULT 0,
        accuracy                REAL,
        precision               REAL,
        recall                  REAL,
        f1_score                REAL,
        mrr                     REAL,
        faithfulness            REAL,
        spearman                REAL,
        created_at              TIMESTAMPTZ DEFAULT now()
    );
    """,
    "CREATE INDEX IF NOT EXISTS idx_system_eval_video_eval_id ON system_eval_video_results (eval_id);",
    "CREATE INDEX IF NOT EXISTS idx_system_eval_evaluated_at ON system_eval (evaluated_at DESC);",
]


def run_migrations():
    """Run all DB migrations. Call once at startup or via CLI."""
    import sys
    raw_url = os.getenv("SUPABASE_URL", "") or os.getenv("DATABASE_URL", "")
    if not raw_url:
        print("ERROR: SUPABASE_URL / DATABASE_URL not set in .env")
        sys.exit(1)
    url = raw_url
    conn = psycopg2.connect(url)
    try:
        with conn.cursor() as cur:
            for i, sql in enumerate(MIGRATIONS, 1):
                print(f"Running migration {i}/{len(MIGRATIONS)}...")
                cur.execute(sql)
        conn.commit()
        print("All migrations completed successfully.")
    except Exception as e:
        conn.rollback()
        print(f"Migration failed: {e}")
        sys.exit(1)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Evaluation helpers
# ---------------------------------------------------------------------------

def save_evaluation_results(run_id: int, evaluation: dict, file_name: str = "unknown") -> None:
    """Insert per-file evaluation into the evaluation_results table."""
    conn = get_db_connection()
    try:
        ct = evaluation.get("check_totals", {})
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO evaluation_results
                    (run_id, file_name, total_rules,
                     text_presence_passed, text_presence_total,
                     page_accuracy_passed, page_accuracy_total,
                     heading_accuracy_passed, heading_accuracy_total,
                     category_validity_passed, category_validity_total,
                     severity_consistency_passed, severity_consistency_total,
                     hallucination_rate, correctness_score, overall_accuracy,
                     failed_rules)
                VALUES (%s,%s,%s, %s,%s, %s,%s, %s,%s, %s,%s, %s,%s, %s,%s,%s, %s)
                """,
                (
                    run_id, file_name, evaluation.get("total_rules", 0),
                    ct.get("text_presence", {}).get("passed", 0),
                    ct.get("text_presence", {}).get("total", 0),
                    ct.get("page_accuracy", {}).get("passed", 0),
                    ct.get("page_accuracy", {}).get("total", 0),
                    ct.get("heading_accuracy", {}).get("passed", 0),
                    ct.get("heading_accuracy", {}).get("total", 0),
                    ct.get("category_validity", {}).get("passed", 0),
                    ct.get("category_validity", {}).get("total", 0),
                    ct.get("severity_consistency", {}).get("passed", 0),
                    ct.get("severity_consistency", {}).get("total", 0),
                    evaluation.get("hallucination_rate"),
                    evaluation.get("correctness_score"),
                    evaluation.get("overall_accuracy"),
                    json.dumps(evaluation.get("failed_rules", [])[:50]),
                ),
            )
        conn.commit()
        # Also store summary in extraction_runs for backward compat
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE extraction_runs SET evaluation_results = %s WHERE id = %s",
                (json.dumps(evaluation), run_id),
            )
        conn.commit()
    finally:
        conn.close()


def _compute_correctness_score(rules: list[dict]) -> float:
    """Compute average cosine similarity between original_text and actionable_rule.

    Uses sentence-transformers (already a dependency) to encode both texts
    and compute cosine similarity.
    """
    try:
        from sentence_transformers import SentenceTransformer
        import numpy as np

        originals = [(r.get("original_text") or "").strip() for r in rules]
        actionables = [(r.get("actionable_rule") or "").strip() for r in rules]

        # Filter pairs where both sides are non-empty
        pairs = [(o, a) for o, a in zip(originals, actionables) if o and a]
        if not pairs:
            return 0.0

        model = SentenceTransformer("all-MiniLM-L6-v2")
        orig_embs = model.encode([p[0] for p in pairs], show_progress_bar=False)
        act_embs = model.encode([p[1] for p in pairs], show_progress_bar=False)

        sims = []
        for oe, ae in zip(orig_embs, act_embs):
            oe_norm = np.linalg.norm(oe)
            ae_norm = np.linalg.norm(ae)
            if oe_norm > 0 and ae_norm > 0:
                sims.append(float(np.dot(oe, ae) / (oe_norm * ae_norm)))

        return round(sum(sims) / len(sims) * 100, 1) if sims else 0.0
    except Exception as exc:
        logger.warning("Correctness score computation failed: %s", exc)
        return 0.0


def run_brutal_evaluation(pdf_path: str, extraction_data: dict) -> dict:
    """5-check hallucination evaluation against the source PDF.

    Checks:
      1. text_presence       — Is original_text found anywhere in the PDF?
      2. page_accuracy       — Is original_text on the claimed page (±1)?
      3. heading_accuracy    — Does section_heading appear on the claimed page?
      4. category_validity   — Are all categories in the allowed set?
      5. severity_consistency — Is severity appropriate for hazard keywords?

    Also computes:
      - hallucination_rate: % of rules failing text/page/heading checks
      - correctness_score:  avg cosine similarity original_text ↔ actionable_rule
    """
    try:
        import pymupdf as fitz  # PyMuPDF (preferred)
    except Exception:
        import fitz  # type: ignore  # PyMuPDF fallback

    rules = extraction_data.get("rules", [])
    if not rules:
        return {"total_rules": 0, "overall_accuracy": 100.0, "checks": {},
                "check_totals": {}, "hallucination_rate": 0.0, "correctness_score": 100.0}

    doc = fitz.open(pdf_path)
    page_texts: dict[int, str] = {}
    for page_num in range(len(doc)):
        page = doc[page_num]
        page_texts[page_num + 1] = page.get_text().strip().lower()
    doc.close()

    ALLOWED_CATEGORIES = {
        "electrical", "chemical", "woodworking", "power_tools",
        "heat_fire", "mechanical", "PPE_required", "child_safety",
        "toxic_exposure", "ventilation", "structural", "general_safety",
    }
    HAZARD_KEYWORDS = [
        "toxic", "fatal", "death", "electrocution", "fire",
        "explosion", "asbestos", "cyanide", "carbon monoxide",
        "burn", "amputation", "crush",
    ]

    results_per_rule = []
    check_totals = {
        "text_presence": {"passed": 0, "total": 0},
        "page_accuracy": {"passed": 0, "total": 0},
        "heading_accuracy": {"passed": 0, "total": 0},
        "category_validity": {"passed": 0, "total": 0},
        "severity_consistency": {"passed": 0, "total": 0},
    }
    hallucination_fail_count = 0  # rules failing text/page/heading

    for rule in rules:
        original_text = (rule.get("original_text") or "").lower().strip()
        page_num = rule.get("page_number")
        section_heading = (rule.get("section_heading") or "").lower().strip()
        actionable = (rule.get("actionable_rule") or "").strip()
        categories = rule.get("categories", [])
        suggested_sev = rule.get("suggested_severity") or 1
        validated_sev = rule.get("validated_severity") or suggested_sev

        checks = {}
        failed = []

        # 1. Text Presence
        text_found = False
        if original_text and len(original_text) > 10:
            for pt in page_texts.values():
                if original_text in pt:
                    text_found = True
                    break
            if not text_found:
                words = original_text.split()
                for pt in page_texts.values():
                    matched_words = sum(1 for w in words if w in pt)
                    if len(words) > 0 and matched_words / len(words) >= 0.6:
                        text_found = True
                        break
        elif original_text:
            text_found = True

        checks["text_presence"] = text_found
        check_totals["text_presence"]["total"] += 1
        if text_found:
            check_totals["text_presence"]["passed"] += 1
        else:
            failed.append("text_presence")

        # 2. Page Accuracy
        page_ok = False
        if page_num and original_text and len(original_text) > 10:
            words = original_text.split()[:8]
            search_str = " ".join(words)
            for offset in [0, -1, 1]:
                check_page = page_num + offset
                if check_page in page_texts and search_str in page_texts[check_page]:
                    page_ok = True
                    break
            if not page_ok:
                for offset in [0, -1, 1]:
                    check_page = page_num + offset
                    if check_page in page_texts:
                        matched = sum(1 for w in words if w in page_texts[check_page])
                        if len(words) > 0 and matched / len(words) >= 0.7:
                            page_ok = True
                            break
        else:
            page_ok = True

        checks["page_accuracy"] = page_ok
        check_totals["page_accuracy"]["total"] += 1
        if page_ok:
            check_totals["page_accuracy"]["passed"] += 1
        else:
            failed.append("page_accuracy")

        # 3. Heading Accuracy — section_heading should appear on claimed page
        heading_ok = True
        if section_heading and page_num and section_heading != "unknown section":
            heading_found = False
            heading_words = section_heading.split()
            for offset in [0, -1, 1]:
                check_page = page_num + offset
                if check_page in page_texts:
                    pt = page_texts[check_page]
                    if section_heading in pt:
                        heading_found = True
                        break
                    # Fuzzy: at least 70% words match
                    if heading_words:
                        matched = sum(1 for w in heading_words if w in pt)
                        if matched / len(heading_words) >= 0.7:
                            heading_found = True
                            break
            heading_ok = heading_found

        checks["heading_accuracy"] = heading_ok
        check_totals["heading_accuracy"]["total"] += 1
        if heading_ok:
            check_totals["heading_accuracy"]["passed"] += 1
        else:
            failed.append("heading_accuracy")

        # 4. Category Validity
        cats_valid = all(c in ALLOWED_CATEGORIES for c in categories) if categories else True
        checks["category_validity"] = cats_valid
        check_totals["category_validity"]["total"] += 1
        if cats_valid:
            check_totals["category_validity"]["passed"] += 1
        else:
            failed.append("category_validity")

        # 5. Severity Consistency
        combined_text = (original_text + " " + actionable.lower())
        has_hazard = any(kw in combined_text for kw in HAZARD_KEYWORDS)
        severity_ok = True
        if has_hazard and validated_sev < 3:
            severity_ok = False
        if validated_sev < suggested_sev:
            severity_ok = False

        checks["severity_consistency"] = severity_ok
        check_totals["severity_consistency"]["total"] += 1
        if severity_ok:
            check_totals["severity_consistency"]["passed"] += 1
        else:
            failed.append("severity_consistency")

        # Track hallucination (text/page/heading failures)
        if not text_found or not page_ok or not heading_ok:
            hallucination_fail_count += 1

        results_per_rule.append({
            "rule_id": rule.get("rule_id", ""),
            "actionable_rule": actionable[:100],
            "checks": checks,
            "all_passed": len(failed) == 0,
            "failed_checks": failed,
        })

    total_rules = len(rules)
    total_checks = sum(ct["total"] for ct in check_totals.values())
    total_passed = sum(ct["passed"] for ct in check_totals.values())

    per_check_accuracy = {
        check_name: round(ct["passed"] / ct["total"] * 100, 1) if ct["total"] > 0 else 100.0
        for check_name, ct in check_totals.items()
    }
    overall_accuracy = round(total_passed / total_checks * 100, 1) if total_checks > 0 else 100.0
    failed_rules = [r for r in results_per_rule if not r["all_passed"]]

    hallucination_rate = round(hallucination_fail_count / total_rules * 100, 1) if total_rules > 0 else 0.0
    correctness_score = _compute_correctness_score(rules)

    return {
        "total_rules": total_rules,
        "total_checks": total_checks,
        "checks_passed": total_passed,
        "overall_accuracy": overall_accuracy,
        "per_check_accuracy": per_check_accuracy,
        "check_totals": check_totals,
        "rules_all_passed": total_rules - len(failed_rules),
        "rules_with_failures": len(failed_rules),
        "hallucination_rate": hallucination_rate,
        "correctness_score": correctness_score,
        "failed_rules": failed_rules[:50],
    }


def run_structure_evaluation(extraction_data: dict) -> dict:
    """Structural evaluation (no PDF needed): rule structure, categories, severity."""
    rules = extraction_data.get("rules", [])
    if not rules:
        return {"total_rules": 0, "overall_accuracy": 100.0, "checks": {}}

    ALLOWED_CATEGORIES = {
        "electrical", "chemical", "woodworking", "power_tools",
        "heat_fire", "mechanical", "PPE_required", "child_safety",
        "toxic_exposure", "ventilation", "structural", "general_safety",
    }

    check_totals = {
        "has_actionable_rule": {"passed": 0, "total": 0},
        "has_original_text": {"passed": 0, "total": 0},
        "category_validity": {"passed": 0, "total": 0},
        "has_severity": {"passed": 0, "total": 0},
    }
    results_per_rule = []

    for rule in rules:
        checks = {}
        failed = []

        has_action = bool((rule.get("actionable_rule") or "").strip())
        checks["has_actionable_rule"] = has_action
        check_totals["has_actionable_rule"]["total"] += 1
        if has_action:
            check_totals["has_actionable_rule"]["passed"] += 1
        else:
            failed.append("has_actionable_rule")

        has_orig = bool((rule.get("original_text") or "").strip())
        checks["has_original_text"] = has_orig
        check_totals["has_original_text"]["total"] += 1
        if has_orig:
            check_totals["has_original_text"]["passed"] += 1
        else:
            failed.append("has_original_text")

        categories = rule.get("categories", [])
        cats_valid = all(c in ALLOWED_CATEGORIES for c in categories) if categories else True
        checks["category_validity"] = cats_valid
        check_totals["category_validity"]["total"] += 1
        if cats_valid:
            check_totals["category_validity"]["passed"] += 1
        else:
            failed.append("category_validity")

        has_sev = rule.get("validated_severity") is not None
        checks["has_severity"] = has_sev
        check_totals["has_severity"]["total"] += 1
        if has_sev:
            check_totals["has_severity"]["passed"] += 1
        else:
            failed.append("has_severity")

        results_per_rule.append({
            "rule_id": rule.get("rule_id", ""),
            "actionable_rule": (rule.get("actionable_rule") or "")[:100],
            "checks": checks,
            "all_passed": len(failed) == 0,
            "failed_checks": failed,
        })

    total_rules = len(rules)
    total_checks = sum(ct["total"] for ct in check_totals.values())
    total_passed = sum(ct["passed"] for ct in check_totals.values())
    per_check_accuracy = {
        check_name: round(ct["passed"] / ct["total"] * 100, 1) if ct["total"] > 0 else 100.0
        for check_name, ct in check_totals.items()
    }
    overall_accuracy = round(total_passed / total_checks * 100, 1) if total_checks > 0 else 100.0
    failed_rules = [r for r in results_per_rule if not r["all_passed"]]

    return {
        "total_rules": total_rules,
        "total_checks": total_checks,
        "checks_passed": total_passed,
        "overall_accuracy": overall_accuracy,
        "per_check_accuracy": per_check_accuracy,
        "rules_all_passed": total_rules - len(failed_rules),
        "rules_with_failures": len(failed_rules),
        "failed_rules": failed_rules[:50],
    }


def _coerce_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


def _tokenize(text: str) -> set[str]:
    if not text:
        return set()
    cleaned = re.sub(r"[^a-z0-9\s]", " ", text.lower())
    words = [w for w in cleaned.split() if len(w) > 2]
    return set(words)


def _average_ranks(values: list[float]) -> list[float]:
    indexed = sorted(enumerate(values), key=lambda x: x[1])
    ranks = [0.0] * len(values)
    i = 0
    while i < len(indexed):
        j = i
        while j + 1 < len(indexed) and indexed[j + 1][1] == indexed[i][1]:
            j += 1
        avg_rank = (i + j + 2) / 2.0
        for k in range(i, j + 1):
            orig_idx = indexed[k][0]
            ranks[orig_idx] = avg_rank
        i = j + 1
    return ranks


def _pearson_corr(xs: list[float], ys: list[float]) -> float | None:
    n = len(xs)
    if n < 2:
        return None
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    num = 0.0
    den_x = 0.0
    den_y = 0.0
    for x, y in zip(xs, ys):
        dx = x - mean_x
        dy = y - mean_y
        num += dx * dy
        den_x += dx * dx
        den_y += dy * dy
    if den_x <= 0.0 or den_y <= 0.0:
        return None
    return num / math.sqrt(den_x * den_y)


def _spearman_corr(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) != len(ys) or len(xs) < 2:
        return None
    rx = _average_ranks(xs)
    ry = _average_ranks(ys)
    return _pearson_corr(rx, ry)


def _binary_metrics(tp: int, tn: int, fp: int, fn: int) -> tuple[float, float, float, float]:
    total = tp + tn + fp + fn
    accuracy = ((tp + tn) / total * 100.0) if total > 0 else 0.0
    precision = (tp / (tp + fp) * 100.0) if (tp + fp) > 0 else 0.0
    recall = (tp / (tp + fn) * 100.0) if (tp + fn) > 0 else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0.0
    return round(accuracy, 2), round(precision, 2), round(recall, 2), round(f1, 2)


YOUTUBE_URL_RE = re.compile(r'https?://(?:www\.)?(?:youtube\.com|youtu\.be)/[^\s,;\]\)>"\']+', re.IGNORECASE)


def _extract_video_id_from_url(value: str) -> str | None:
    raw = (value or "").strip()
    if not raw:
        return None
    # Allow direct video-id style strings.
    if re.fullmatch(r"[A-Za-z0-9_-]{11}", raw):
        return raw

    try:
        p = urlparse(raw)
        host = p.netloc.lower()
        path = p.path.strip("/")
        if "youtu.be" in host:
            vid = path.split("/")[0]
            return vid if re.fullmatch(r"[A-Za-z0-9_-]{11}", vid or "") else None
        if "youtube.com" in host:
            qs = parse_qs(p.query)
            vid = (qs.get("v") or [None])[0]
            if vid and re.fullmatch(r"[A-Za-z0-9_-]{11}", vid):
                return vid
            # Support /shorts/{id}
            if path.startswith("shorts/"):
                vid = path.split("/")[1] if len(path.split("/")) > 1 else None
                return vid if re.fullmatch(r"[A-Za-z0-9_-]{11}", vid or "") else None
    except Exception:
        return None
    return None


def _normalize_youtube_url(value: str) -> str | None:
    vid = _extract_video_id_from_url(value)
    if not vid:
        return None
    return f"https://www.youtube.com/watch?v={vid}"


def _extract_urls_from_text_blob(text: str) -> list[str]:
    if not text:
        return []
    found = YOUTUBE_URL_RE.findall(text)
    # Also support comma/space-separated raw ids.
    tokens = re.split(r"[\s,;]+", text)
    for token in tokens:
        if re.fullmatch(r"[A-Za-z0-9_-]{11}", token or ""):
            found.append(token)
    cleaned: list[str] = []
    seen = set()
    for item in found:
        norm = _normalize_youtube_url(item)
        if norm and norm not in seen:
            cleaned.append(norm)
            seen.add(norm)
    return cleaned


def _insert_urls_into_bucket(urls: list[str], source_type: str, source_file: str | None = None) -> tuple[int, int]:
    if not urls:
        return 0, 0
    conn = get_db_connection()
    inserted = 0
    try:
        with conn.cursor() as cur:
            for url in urls:
                vid = _extract_video_id_from_url(url)
                cur.execute(
                    """
                    INSERT INTO youtube_urls (url, video_id, source_type, source_file, last_used_at)
                    VALUES (%s, %s, %s, %s, now())
                    ON CONFLICT (url)
                    DO UPDATE SET last_used_at = now(), source_type = EXCLUDED.source_type, source_file = EXCLUDED.source_file
                    """,
                    (url, vid, source_type, source_file),
                )
                if cur.rowcount > 0:
                    inserted += 1
        conn.commit()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM youtube_urls")
            total = int(cur.fetchone()[0])
        return inserted, total
    finally:
        conn.close()


def _get_url_pool() -> list[str]:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT url FROM youtube_urls ORDER BY created_at ASC")
            return [str(r[0]) for r in cur.fetchall()]
    finally:
        conn.close()


def _evaluate_system(limit: int = 50, youtube_urls: list[str] | None = None, random_count: int | None = None) -> dict:
    """Evaluate completed scans, optionally filtered by provided YouTube URLs."""
    try:
        from rule_extraction.extract_rules import override_severity
    except ModuleNotFoundError:
        import sys
        project_root = Path(__file__).resolve().parent.parent
        root_str = str(project_root)
        if root_str not in sys.path:
            sys.path.insert(0, root_str)
        from rule_extraction.extract_rules import override_severity

    conn = get_db_connection()
    try:
        selected_urls = [_normalize_youtube_url(u) for u in (youtube_urls or [])]
        selected_urls = [u for u in selected_urls if u]
        selected_urls = list(dict.fromkeys(selected_urls))

        if random_count is not None and selected_urls:
            n = max(1, min(int(random_count), len(selected_urls)))
            selected_urls = random.sample(selected_urls, n)

        selected_video_ids = [v for v in (_extract_video_id_from_url(u) for u in selected_urls) if v]
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if selected_video_ids:
                cur.execute(
                    """
                    SELECT DISTINCT ON (video_id) id, video_id, video_url, title, scan_timestamp, output_json
                    FROM completed_scans
                    WHERE output_json IS NOT NULL AND video_id = ANY(%s)
                    ORDER BY video_id, scan_timestamp DESC
                    """,
                    (selected_video_ids,),
                )
                scan_rows = [dict(r) for r in cur.fetchall()]
            else:
                cur.execute(
                    """
                    SELECT id, video_id, video_url, title, scan_timestamp, output_json
                    FROM completed_scans
                    WHERE output_json IS NOT NULL
                    ORDER BY scan_timestamp DESC
                    LIMIT %s
                    """,
                    (limit,),
                )
                scan_rows = [dict(r) for r in cur.fetchall()]

        if selected_video_ids:
            scanned_ids = {str(s.get("video_id")) for s in scan_rows if s.get("video_id")}
            missing_urls = [u for u in selected_urls if (_extract_video_id_from_url(u) not in scanned_ids)]
        else:
            missing_urls = []

        tp = tn = fp = fn = 0
        all_llm_scores: list[float] = []
        all_override_scores: list[float] = []
        reciprocal_ranks: list[float] = []
        total_precautions = 0
        supported_precautions = 0
        total_steps = 0
        evaluated_scans = 0
        scan_breakdown: list[dict[str, Any]] = []
        per_video_rows: list[dict[str, Any]] = []

        for scan in scan_rows:
            output_json = scan.get("output_json") or {}
            if isinstance(output_json, str):
                try:
                    output_json = json.loads(output_json)
                except Exception:
                    output_json = {}

            steps = output_json.get("steps") or []
            if not isinstance(steps, list) or not steps:
                continue

            model_reports = output_json.get("modelReports") or {}
            qwen_report = model_reports.get("qwen") if isinstance(model_reports, dict) else None
            if not qwen_report:
                qwen_report = output_json.get("report")
            if not isinstance(qwen_report, dict):
                continue

            step_analysis = qwen_report.get("step_safety_analysis") or []
            if not isinstance(step_analysis, list):
                step_analysis = []
            by_step: dict[int, dict[str, Any]] = {}
            for s in step_analysis:
                if isinstance(s, dict):
                    num = s.get("step_number")
                    if isinstance(num, int):
                        by_step[num] = s

            tmp_rules = []
            for idx, step in enumerate(steps, start=1):
                if not isinstance(step, dict):
                    continue
                step_num = step.get("step_number")
                if not isinstance(step_num, int):
                    step_num = idx
                step_text = (step.get("step_text") or "").strip()
                action_summary = (step.get("action_summary") or step_text).strip()
                transcript_excerpt = (step.get("transcript_excerpt") or step_text).strip()
                tmp_rules.append(
                    {
                        "rule_id": str(step_num),
                        "suggested_severity": 1,
                        "validated_severity": 1,
                        "actionable_rule": action_summary,
                        "original_text": transcript_excerpt,
                    }
                )

            scored_steps = override_severity(tmp_rules)
            override_by_step = {
                int(r.get("rule_id")): int(r.get("validated_severity") or 1)
                for r in scored_steps
                if str(r.get("rule_id", "")).isdigit()
            }

            scan_llm: list[float] = []
            scan_override: list[float] = []
            scan_rrs: list[float] = []
            scan_total_prec = 0
            scan_supported_prec = 0
            scan_tp = scan_tn = scan_fp = scan_fn = 0

            for idx, step in enumerate(steps, start=1):
                if not isinstance(step, dict):
                    continue
                step_num = step.get("step_number")
                if not isinstance(step_num, int):
                    step_num = idx
                analysis = by_step.get(step_num, {})
                llm_score = _coerce_float(analysis.get("risk_level"), default=float("nan"))
                if math.isnan(llm_score):
                    continue

                override_score = float(override_by_step.get(step_num, 1))
                total_steps += 1

                scan_llm.append(llm_score)
                scan_override.append(override_score)
                all_llm_scores.append(llm_score)
                all_override_scores.append(override_score)

                y_pred = 1 if llm_score >= 3.0 else 0
                y_true = 1 if override_score >= 3.0 else 0
                if y_true == 1 and y_pred == 1:
                    tp += 1
                    scan_tp += 1
                elif y_true == 0 and y_pred == 0:
                    tn += 1
                    scan_tn += 1
                elif y_true == 0 and y_pred == 1:
                    fp += 1
                    scan_fp += 1
                else:
                    fn += 1
                    scan_fn += 1

                matched_rules = analysis.get("matched_rules") or []
                rr = 0.0
                if isinstance(matched_rules, list) and matched_rules:
                    for rank, rule in enumerate(matched_rules, start=1):
                        if not isinstance(rule, dict):
                            continue
                        sev = _coerce_float(rule.get("severity"), default=0.0)
                        if sev >= max(3.0, override_score):
                            rr = 1.0 / rank
                            break
                reciprocal_ranks.append(rr)
                scan_rrs.append(rr)

                required_precautions = analysis.get("required_precautions") or []
                if isinstance(required_precautions, list):
                    evidence_text = " ".join(
                        [
                            str(step.get("step_text") or ""),
                            str(step.get("transcript_excerpt") or ""),
                            " ".join(
                                str((r.get("rule_text") or r.get("actionable_rule") or ""))
                                for r in matched_rules if isinstance(r, dict)
                            ),
                        ]
                    )
                    evidence_tokens = _tokenize(evidence_text)
                    for precaution in required_precautions:
                        prec_tokens = _tokenize(str(precaution))
                        if not prec_tokens:
                            continue
                        scan_total_prec += 1
                        overlap = len(prec_tokens & evidence_tokens)
                        support_ratio = overlap / len(prec_tokens)
                        if overlap >= 2 and support_ratio >= 0.5:
                            scan_supported_prec += 1

            if scan_llm:
                evaluated_scans += 1
                total_precautions += scan_total_prec
                supported_precautions += scan_supported_prec
                scan_spearman = _spearman_corr(scan_llm, scan_override)
                scan_acc, scan_prec, scan_rec, scan_f1 = _binary_metrics(scan_tp, scan_tn, scan_fp, scan_fn)
                scan_mrr = round(sum(scan_rrs) / len(scan_rrs), 4) if scan_rrs else 0.0
                scan_faith = round((scan_supported_prec / scan_total_prec) * 100.0, 2) if scan_total_prec > 0 else 100.0
                scan_breakdown.append(
                    {
                        "scan_id": scan.get("id"),
                        "video_id": scan.get("video_id"),
                        "title": scan.get("title"),
                        "scan_timestamp": scan.get("scan_timestamp").isoformat() if scan.get("scan_timestamp") else None,
                        "steps_evaluated": len(scan_llm),
                        "avg_llm_risk": round(sum(scan_llm) / len(scan_llm), 3),
                        "avg_override_risk": round(sum(scan_override) / len(scan_override), 3),
                        "scan_spearman": round(scan_spearman, 4) if scan_spearman is not None else None,
                        "scan_mrr": scan_mrr,
                        "faithfulness": scan_faith,
                    }
                )
                per_video_rows.append(
                    {
                        "video_id": scan.get("video_id"),
                        "video_url": scan.get("video_url") or (f"https://www.youtube.com/watch?v={scan.get('video_id')}" if scan.get("video_id") else None),
                        "scan_id": scan.get("id"),
                        "steps_evaluated": len(scan_llm),
                        "total_precautions": scan_total_prec,
                        "supported_precautions": scan_supported_prec,
                        "tp": scan_tp,
                        "tn": scan_tn,
                        "fp": scan_fp,
                        "fn": scan_fn,
                        "accuracy": scan_acc,
                        "precision": scan_prec,
                        "recall": scan_rec,
                        "f1_score": scan_f1,
                        "mrr": scan_mrr,
                        "faithfulness": scan_faith,
                        "spearman": round(scan_spearman, 4) if scan_spearman is not None else None,
                    }
                )

        accuracy, precision, recall, f1_score = _binary_metrics(tp, tn, fp, fn)
        mrr = round(sum(reciprocal_ranks) / len(reciprocal_ranks), 4) if reciprocal_ranks else 0.0
        faithfulness = round((supported_precautions / total_precautions) * 100.0, 2) if total_precautions > 0 else 100.0
        spearman = _spearman_corr(all_llm_scores, all_override_scores)

        details_json = {
            "notes": {
                "accuracy": "Step-level agreement of unsafe-vs-safe labels (threshold >=3).",
                "precision": "Among LLM-predicted unsafe steps, how many are unsafe by override severity.",
                "recall": "Among override-unsafe steps, how many are predicted unsafe by LLM.",
                "f1_score": "Harmonic mean of precision and recall.",
                "mean_reciprocal_rank": "Mean reciprocal rank of first high-severity matched rule per step.",
                "faithfulness_score": "Percent of required precautions supported by step evidence and matched rules.",
                "spearman_correlation": "Rank correlation between Qwen step risk levels and override severity scores.",
            },
            "scan_breakdown": scan_breakdown,
            "selected_urls": selected_urls,
            "missing_urls": missing_urls,
            "llm_override_pairs": [
                {"llm": round(l, 3), "override": round(o, 3)}
                for l, o in zip(all_llm_scores[:300], all_override_scores[:300])
            ],
        }

        result = {
            "evaluated_at": datetime.now(timezone.utc).isoformat(),
            "model_key": "qwen",
            "sample_size": limit,
            "youtube_urls": selected_urls,
            "selected_urls_count": len(selected_urls) if selected_urls else len(scan_rows),
            "evaluated_scans": evaluated_scans,
            "total_steps": total_steps,
            "total_precautions": total_precautions,
            "supported_precautions": supported_precautions,
            "confusion_matrix": {
                "true_positive": tp,
                "true_negative": tn,
                "false_positive": fp,
                "false_negative": fn,
            },
            "metrics": {
                "accuracy": accuracy,
                "precision": precision,
                "recall": recall,
                "f1_score": f1_score,
                "mean_reciprocal_rank": mrr,
                "faithfulness_score": faithfulness,
                "spearman_correlation": round(spearman, 4) if spearman is not None else None,
            },
            "details": details_json,
            "missing_urls": missing_urls,
        }

        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO system_eval
                    (model_key, sample_size, evaluated_scans, total_steps,
                     total_precautions, supported_precautions,
                     true_positive, true_negative, false_positive, false_negative,
                     accuracy, precision, recall, f1_score,
                     mean_reciprocal_rank, faithfulness_score, spearman_correlation,
                     details_json, youtube_urls, selected_urls_count, total_urls_in_pool)
                VALUES (%s, %s, %s, %s,
                        %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s,
                    %s, %s, %s, %s)
                RETURNING id, evaluated_at
                """,
                (
                    "qwen",
                    limit,
                    evaluated_scans,
                    total_steps,
                    total_precautions,
                    supported_precautions,
                    tp,
                    tn,
                    fp,
                    fn,
                    accuracy,
                    precision,
                    recall,
                    f1_score,
                    mrr,
                    faithfulness,
                    (round(spearman, 4) if spearman is not None else None),
                    psycopg2.extras.Json(details_json),
                    psycopg2.extras.Json(selected_urls),
                    (len(selected_urls) if selected_urls else len(scan_rows)),
                    len(_get_url_pool()),
                ),
            )
            inserted = cur.fetchone()

            eval_id = inserted["id"] if inserted else None
            if eval_id:
                for row in per_video_rows:
                    cur.execute(
                        """
                        INSERT INTO system_eval_video_results
                            (eval_id, video_id, video_url, scan_id, steps_evaluated,
                             total_precautions, supported_precautions,
                             true_positive, true_negative, false_positive, false_negative,
                             accuracy, precision, recall, f1_score, mrr, faithfulness, spearman)
                        VALUES (%s, %s, %s, %s, %s,
                                %s, %s,
                                %s, %s, %s, %s,
                                %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            eval_id,
                            row.get("video_id"),
                            row.get("video_url"),
                            row.get("scan_id"),
                            row.get("steps_evaluated"),
                            row.get("total_precautions"),
                            row.get("supported_precautions"),
                            row.get("tp"),
                            row.get("tn"),
                            row.get("fp"),
                            row.get("fn"),
                            row.get("accuracy"),
                            row.get("precision"),
                            row.get("recall"),
                            row.get("f1_score"),
                            row.get("mrr"),
                            row.get("faithfulness"),
                            row.get("spearman"),
                        ),
                    )

                cur.execute(
                    """
                    SELECT
                        COALESCE(SUM(steps_evaluated), 0) AS total_steps,
                        COALESCE(SUM(total_precautions), 0) AS total_precautions,
                        COALESCE(SUM(supported_precautions), 0) AS supported_precautions,
                        COALESCE(SUM(true_positive), 0) AS tp,
                        COALESCE(SUM(true_negative), 0) AS tn,
                        COALESCE(SUM(false_positive), 0) AS fp,
                        COALESCE(SUM(false_negative), 0) AS fn
                    FROM system_eval_video_results
                    """
                )
                agg = cur.fetchone() or {}
                cum_acc, cum_prec, cum_rec, cum_f1 = _binary_metrics(
                    int(agg.get("tp") or 0),
                    int(agg.get("tn") or 0),
                    int(agg.get("fp") or 0),
                    int(agg.get("fn") or 0),
                )
                cur.execute(
                    """
                    UPDATE system_eval
                    SET cum_total_steps = %s,
                        cum_total_precautions = %s,
                        cum_supported_precautions = %s,
                        cum_true_positive = %s,
                        cum_true_negative = %s,
                        cum_false_positive = %s,
                        cum_false_negative = %s,
                        cum_accuracy = %s,
                        cum_precision = %s,
                        cum_recall = %s,
                        cum_f1_score = %s
                    WHERE id = %s
                    """,
                    (
                        int(agg.get("total_steps") or 0),
                        int(agg.get("total_precautions") or 0),
                        int(agg.get("supported_precautions") or 0),
                        int(agg.get("tp") or 0),
                        int(agg.get("tn") or 0),
                        int(agg.get("fp") or 0),
                        int(agg.get("fn") or 0),
                        cum_acc,
                        cum_prec,
                        cum_rec,
                        cum_f1,
                        eval_id,
                    ),
                )

            conn.commit()

        result["id"] = inserted["id"] if inserted else None
        if inserted and inserted.get("evaluated_at"):
            result["evaluated_at"] = inserted["evaluated_at"].isoformat()
        if inserted:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT cum_total_steps, cum_total_precautions, cum_supported_precautions,
                           cum_true_positive, cum_true_negative, cum_false_positive, cum_false_negative,
                           cum_accuracy, cum_precision, cum_recall, cum_f1_score
                    FROM system_eval
                    WHERE id = %s
                    """,
                    (inserted["id"],),
                )
                c = cur.fetchone() or {}
            result["cumulative"] = {
                "total_steps": int(c.get("cum_total_steps") or 0),
                "total_precautions": int(c.get("cum_total_precautions") or 0),
                "supported_precautions": int(c.get("cum_supported_precautions") or 0),
                "confusion_matrix": {
                    "true_positive": int(c.get("cum_true_positive") or 0),
                    "true_negative": int(c.get("cum_true_negative") or 0),
                    "false_positive": int(c.get("cum_false_positive") or 0),
                    "false_negative": int(c.get("cum_false_negative") or 0),
                },
                "metrics": {
                    "accuracy": _coerce_float(c.get("cum_accuracy"), 0.0),
                    "precision": _coerce_float(c.get("cum_precision"), 0.0),
                    "recall": _coerce_float(c.get("cum_recall"), 0.0),
                    "f1_score": _coerce_float(c.get("cum_f1_score"), 0.0),
                },
            }
        return result
    finally:
        conn.close()


def _fetch_latest_system_evaluation() -> dict | None:
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, evaluated_at, model_key, sample_size, evaluated_scans,
                       total_steps, total_precautions, supported_precautions,
                       true_positive, true_negative, false_positive, false_negative,
                       accuracy, precision, recall, f1_score,
                       mean_reciprocal_rank, faithfulness_score, spearman_correlation,
                      details_json, youtube_urls, selected_urls_count, total_urls_in_pool,
                      cum_total_steps, cum_total_precautions, cum_supported_precautions,
                      cum_true_positive, cum_true_negative, cum_false_positive, cum_false_negative,
                      cum_accuracy, cum_precision, cum_recall, cum_f1_score
                FROM system_eval
                ORDER BY evaluated_at DESC
                LIMIT 1
                """
            )
            row = cur.fetchone()
            if not row:
                return None

        result = dict(row)
        if result.get("evaluated_at"):
            result["evaluated_at"] = result["evaluated_at"].isoformat()
        details = result.get("details_json") or {}
        if isinstance(details, str):
            try:
                details = json.loads(details)
            except Exception:
                details = {}

        return {
            "id": result.get("id"),
            "evaluated_at": result.get("evaluated_at"),
            "model_key": result.get("model_key"),
            "sample_size": result.get("sample_size"),
            "youtube_urls": result.get("youtube_urls") or [],
            "selected_urls_count": result.get("selected_urls_count") or 0,
            "total_urls_in_pool": result.get("total_urls_in_pool") or 0,
            "evaluated_scans": result.get("evaluated_scans"),
            "total_steps": result.get("total_steps"),
            "total_precautions": result.get("total_precautions"),
            "supported_precautions": result.get("supported_precautions"),
            "confusion_matrix": {
                "true_positive": result.get("true_positive", 0),
                "true_negative": result.get("true_negative", 0),
                "false_positive": result.get("false_positive", 0),
                "false_negative": result.get("false_negative", 0),
            },
            "metrics": {
                "accuracy": _coerce_float(result.get("accuracy"), 0.0),
                "precision": _coerce_float(result.get("precision"), 0.0),
                "recall": _coerce_float(result.get("recall"), 0.0),
                "f1_score": _coerce_float(result.get("f1_score"), 0.0),
                "mean_reciprocal_rank": _coerce_float(result.get("mean_reciprocal_rank"), 0.0),
                "faithfulness_score": _coerce_float(result.get("faithfulness_score"), 0.0),
                "spearman_correlation": result.get("spearman_correlation"),
            },
            "cumulative": {
                "total_steps": int(result.get("cum_total_steps") or 0),
                "total_precautions": int(result.get("cum_total_precautions") or 0),
                "supported_precautions": int(result.get("cum_supported_precautions") or 0),
                "confusion_matrix": {
                    "true_positive": int(result.get("cum_true_positive") or 0),
                    "true_negative": int(result.get("cum_true_negative") or 0),
                    "false_positive": int(result.get("cum_false_positive") or 0),
                    "false_negative": int(result.get("cum_false_negative") or 0),
                },
                "metrics": {
                    "accuracy": _coerce_float(result.get("cum_accuracy"), 0.0),
                    "precision": _coerce_float(result.get("cum_precision"), 0.0),
                    "recall": _coerce_float(result.get("cum_recall"), 0.0),
                    "f1_score": _coerce_float(result.get("cum_f1_score"), 0.0),
                },
            },
            "details": details,
        }
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Extraction pipeline (calls safety-extraction CLI subprocess)
# ---------------------------------------------------------------------------

def _safety_extraction_dir() -> Path:
    return Path(__file__).parent.parent / "rule_extraction"


def _find_python() -> str:
    candidates = ["python", "python3", "py"] if os.name == "nt" else ["python3", "python"]
    for candidate in candidates:
        path = shutil.which(candidate)
        if path:
            return candidate
    return candidates[0]


def _strip_embeddings(data: dict) -> dict:
    if "rules" in data and isinstance(data["rules"], list):
        for rule in data["rules"]:
            if isinstance(rule, dict):
                rule.pop("embedding", None)
    return data


def _get_supabase_project_ref() -> str | None:
    """Extract project ref from SUPABASE_URL (supports both direct and pooler URLs)."""
    url = os.getenv("SUPABASE_URL", "")
    # Direct connection: postgresql://postgres:pass@db.REF.supabase.co:5432/postgres
    m = re.search(r"@db\.([^.]+)\.supabase\.co", url)
    if m:
        return m.group(1)
    # Session pooler: postgresql://postgres.REF:pass@...pooler.supabase.com:...
    m = re.search(r"postgres\.([^:]+):", url)
    if m:
        return m.group(1)
    # SUPABASE_PROJECT_REF env var as explicit override
    return os.getenv("SUPABASE_PROJECT_REF") or None


def _upload_to_supabase_storage(file_path: str, original_filename: str) -> str | None:
    """Upload PDF to Supabase Storage bucket and return public URL."""
    try:
        import urllib.request
        import urllib.error

        project_ref = _get_supabase_project_ref()
        if not project_ref:
            print("WARNING: Could not determine Supabase project ref — file upload skipped.")
            return None

        api_key = (
            os.getenv("SUPABASE_SERVICE_KEY")
            or os.getenv("SUPABASE_ANON_KEY")
            or os.getenv("SUPABASE_KEY")
        )
        if not api_key:
            print("WARNING: No SUPABASE_SERVICE_KEY or SUPABASE_ANON_KEY set — file upload skipped.")
            return None

        bucket = "safety-pdfs"
        # Use timestamp to avoid collisions
        ts = int(time.time())
        safe_name = re.sub(r"[^a-zA-Z0-9._\-]", "_", original_filename)
        storage_path = f"{ts}_{safe_name}"

        storage_url = f"https://{project_ref}.supabase.co/storage/v1/object/{bucket}/{storage_path}"

        with open(file_path, "rb") as f:
            file_bytes = f.read()

        req = urllib.request.Request(
            storage_url,
            data=file_bytes,
            method="POST",
            headers={
                "Authorization": f"Bearer {api_key}",
                "apikey": api_key,
                "Content-Type": "application/pdf",
                "x-upsert": "true",
            },
        )

        try:
            with urllib.request.urlopen(req) as resp:
                status = resp.getcode()
                if status in (200, 201):
                    public_url = f"https://{project_ref}.supabase.co/storage/v1/object/public/{bucket}/{storage_path}"
                    print(f"File uploaded to Supabase Storage: {public_url}")
                    return public_url
                else:
                    print(f"WARNING: Storage upload returned unexpected status {status}")
                    return None
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            print(f"WARNING: Storage upload failed [{e.code}]: {body}")
            return None

    except Exception as exc:
        print(f"WARNING: Storage upload exception — {exc}")
        return None


def _insert_run_and_rules(
    extraction_data: dict, original_filename: str, file_url: str | None
) -> int:
    conn = get_db_connection()
    try:
        source_docs = [original_filename]
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO extraction_runs
                    (run_timestamp, model_used, total_pages, rule_count,
                     document_count, source_documents, json_source_file, file_url)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    extraction_data.get("extraction_timestamp", datetime.now(timezone.utc).isoformat()),
                    extraction_data.get("model_used", "unknown"),
                    extraction_data.get("total_pages", 0),
                    extraction_data.get("rule_count", 0),
                    extraction_data.get("document_count", 1),
                    source_docs,
                    original_filename,
                    file_url,
                ),
            )
            run_id = cur.fetchone()[0]

            rules = extraction_data.get("rules", [])
            for rule in rules:
                emb = rule.get("embedding")
                emb_str = None
                if emb is not None:
                    if hasattr(emb, "tolist"):
                        emb = emb.tolist()
                    emb_str = "[" + ",".join(str(float(v)) for v in emb) + "]"

                cur.execute(
                    """
                    INSERT INTO safety_rules
                        (rule_id, original_text, actionable_rule, materials,
                         suggested_severity, validated_severity, categories,
                         source_document, page_number, section_heading,
                         embedding, run_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (rule_id) DO NOTHING
                    """,
                    (
                        rule.get("rule_id"),
                        rule.get("original_text", ""),
                        rule.get("actionable_rule", ""),
                        rule.get("materials", []),
                        rule.get("suggested_severity"),
                        rule.get("validated_severity"),
                        rule.get("categories", []),
                        original_filename,
                        rule.get("page_number"),
                        rule.get("section_heading", "Unknown Section"),
                        emb_str,
                        run_id,
                    ),
                )

            conn.commit()
            return run_id
    finally:
        conn.close()


def _prepare_env() -> dict:
    env = os.environ.copy()
    env["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    env["USE_TF"] = "0"
    # Force child Python process to use UTF-8 for stdout/stderr on Windows
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    supabase_url = os.getenv("SUPABASE_URL", "")
    if supabase_url and not env.get("DATABASE_URL"):
        env["DATABASE_URL"] = supabase_url
    groq_key = os.getenv("GROQ_API_KEY", "")
    if groq_key:
        env["GROQ_API_KEY"] = groq_key
    return env


def extract_rules_v2(file_path: str, original_filename: str) -> dict:
    input_path = Path(file_path)
    if not input_path.exists():
        raise Exception(f"File not found: {file_path}")
    if input_path.suffix.lower() != ".pdf":
        raise Exception("Only PDF files are supported.")

    stem = input_path.stem
    ts = int(time.time())
    out_path = Path(tempfile.gettempdir()) / f"{stem}_{ts}_rules.json"

    rule_extraction_dir = _safety_extraction_dir()
    if not rule_extraction_dir.exists():
        raise Exception(f"rule_extraction directory not found at {rule_extraction_dir}")

    python = _find_python()
    env = _prepare_env()

    result = subprocess.run(
        [python, "extract_rules.py", str(input_path), "--output", str(out_path)],
        cwd=str(rule_extraction_dir),
        capture_output=True,
        text=True,
        encoding='utf-8',
        env=env,
    )

    if result.returncode != 0:
        raise Exception(
            f"Extraction failed (exit {result.returncode}):\n"
            f"{result.stderr}{result.stdout}"
        )

    raw = out_path.read_text(encoding='utf-8')
    data = json.loads(raw)

    try:
        out_path.unlink()
    except OSError:
        pass

    file_url = _upload_to_supabase_storage(file_path, original_filename)
    run_id = _insert_run_and_rules(data, original_filename, file_url)

    evaluation = run_brutal_evaluation(file_path, data)
    save_evaluation_results(run_id, evaluation, file_name=original_filename)

    _strip_embeddings(data)

    return {
        "extraction": data,
        "run_id": run_id,
        "evaluation_results": evaluation,
    }


def extract_rules_with_progress(
    file_path: str,
    original_filename: str,
    progress_callback,
) -> dict:
    import re as _re

    input_path = Path(file_path)
    if not input_path.exists():
        raise Exception(f"File not found: {file_path}")

    stem = input_path.stem
    ts = int(time.time())
    out_path = Path(tempfile.gettempdir()) / f"{stem}_{ts}_rules.json"

    rule_extraction_dir = _safety_extraction_dir()
    if not rule_extraction_dir.exists():
        raise Exception(f"rule_extraction directory not found at {rule_extraction_dir}")

    python = _find_python()
    env = _prepare_env()

    progress_callback("upload", {"status": f"File received: {original_filename}", "file": original_filename})
    progress_callback("ingestion", {"status": f"Starting pipeline for {original_filename}", "file": original_filename})

    proc = subprocess.Popen(
        [python, "extract_rules.py", str(input_path), "--output", str(out_path)],
        cwd=str(rule_extraction_dir),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding='utf-8',
        env=env,
        bufsize=1,
    )

    re_ingest = _re.compile(r"Ingesting PDF: (.+?) \((\d+) pages\)")
    re_page_content = _re.compile(r"PDF ingestion complete: (\d+) pages with content")
    re_page_extract = _re.compile(r"Page (\d+): extracted (\d+) rules")
    re_pre_dedup = _re.compile(r"Pre-dedup rules for '(.+?)': (\d+)")
    re_loading_embed = _re.compile(r"Loading embedding model")
    re_embed_done = _re.compile(r"Generated embeddings for (\d+) rules")
    re_dedup_done = _re.compile(r"Deduplication: (\d+) rules → (\d+) rules")
    re_pipeline_done = _re.compile(r"Pipeline complete .+?: (\d+) final rules")
    re_batch_done = _re.compile(r"Batch complete: (\d+) final deduplicated")
    re_batch_summary = _re.compile(r"BATCH SUMMARY — (\d+) documents?, (\d+) total")

    total_pages = 0
    current_page = 0
    last_step = "ingestion"
    all_output: list[str] = []

    for line in proc.stdout:
        line = line.rstrip()
        if not line:
            continue
        all_output.append(line)

        m = re_ingest.search(line)
        if m:
            total_pages = int(m.group(2))
            progress_callback("ingestion", {"status": f"Reading PDF: {m.group(1)} ({total_pages} pages)", "pages": total_pages})
            continue

        m = re_page_content.search(line)
        if m:
            progress_callback("ingestion", {"status": f"PDF read: {m.group(1)} pages with content", "pages": int(m.group(1)), "percentage": 100})
            last_step = "llm_extraction"
            continue

        m = re_page_extract.search(line)
        if m:
            current_page = int(m.group(1))
            rules_on_page = int(m.group(2))
            pct = int(current_page / max(total_pages, 1) * 100)
            progress_callback("llm_extraction", {"status": f"Page {current_page}/{total_pages}: {rules_on_page} rules extracted", "current_page": current_page, "total_pages": total_pages, "percentage": pct})
            continue

        if re_pre_dedup.search(line):
            progress_callback("llm_extraction", {"status": line, "percentage": 100})
            last_step = "embedding"
            continue

        if re_loading_embed.search(line):
            progress_callback("embedding", {"status": "Loading embedding model..."})
            continue

        m = re_embed_done.search(line)
        if m:
            progress_callback("embedding", {"status": f"Embeddings generated for {m.group(1)} rules", "percentage": 100})
            last_step = "deduplication"
            continue

        m = re_dedup_done.search(line)
        if m:
            progress_callback("deduplication", {"status": f"Deduplicated: {m.group(1)} → {m.group(2)} rules", "before": int(m.group(1)), "after": int(m.group(2))})
            continue

        m = re_pipeline_done.search(line)
        if m:
            progress_callback("complete", {"status": f"Pipeline done: {m.group(1)} final rules", "rule_count": int(m.group(1))})
            continue

        m = re_batch_done.search(line)
        if m:
            progress_callback("complete", {"status": f"Batch complete: {m.group(1)} rules", "rule_count": int(m.group(1))})
            continue

        if re_batch_summary.search(line):
            progress_callback("complete", {"status": line})
            continue

    proc.wait()
    if proc.returncode != 0:
        error_lines = "\n".join(all_output[-20:])
        raise Exception(f"Extraction subprocess failed (exit {proc.returncode}):\n{error_lines}")

    if not out_path.exists():
        raise Exception("Extraction completed but no output file was produced.")

    raw = out_path.read_text(encoding='utf-8')
    data = json.loads(raw)
    try:
        out_path.unlink()
    except OSError:
        pass

    file_url = _upload_to_supabase_storage(file_path, original_filename)
    run_id = _insert_run_and_rules(data, original_filename, file_url)

    evaluation = run_brutal_evaluation(file_path, data)
    save_evaluation_results(run_id, evaluation, file_name=original_filename)

    _strip_embeddings(data)

    return {
        "extraction": data,
        "run_id": run_id,
        "evaluation_results": evaluation,
    }


# ---------------------------------------------------------------------------
# Safety analysis (LLM-based, via Groq)
# ---------------------------------------------------------------------------

GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"

SAFETY_ANALYSIS_PROMPT = """You are an expert DIY Safety Analyst and Compliance Officer.

You are given:
1. Extracted DIY steps from a video tutorial (with transcript excerpts)
2. Matched safety rules from a professional compliance database (matched via cosine similarity on semantic embeddings)
3. The video's safety-relevant categories

Your task is to produce a THOROUGH, DETAILED safety assessment.

FOR EACH STEP:
- Identify what safety precautions MUST be followed (based on matched rules AND general safety knowledge for the category)
- Identify what safety measures the creator ALREADY mentions or demonstrates in the video
- Identify what safety measures are MISSING and need to be added
- Assign a risk level (1-5 scale):
  1 = Minimal risk, basic caution
  2 = Low risk, some care needed
  3 = Moderate risk, specific precautions required
  4 = High risk, serious injury possible without precautions
  5 = Critical risk, professional supervision recommended

OVERALL ASSESSMENT:
- SAFE: No significant safety gaps, creator follows good practices
- UNSAFE: Important safety precautions are missing or violated
- PROFESSIONAL_REQUIRED: Activity involves hazards that need trained professional oversight

PARENT MONITORING:
Assess whether parent/adult monitoring is needed based on:
- Tools and materials used (sharp objects, chemicals, heat sources, power tools)
- Skill level required
- Potential for serious injury
- Whether the activity is suitable for children/teens without supervision
- This applies to ALL videos regardless of stated audience

CRITICAL: Your response must be ONLY valid JSON. No markdown, no code fences, no explanation text.

OUTPUT FORMAT:
{
  "verdict": "SAFE" | "UNSAFE" | "PROFESSIONAL_REQUIRED",
  "overall_risk_score": <1.0-5.0>,
  "parent_monitoring_required": true | false,
  "parent_monitoring_reason": "<concise reason why monitoring is/isn't needed>",
  "summary": "<2-3 sentence overall safety summary>",
  "critical_concerns": ["<most important concern 1>", "<concern 2>"],
  "step_safety_analysis": [
    {
      "step_number": 1,
      "action_summary": "<what the step does>",
      "risk_level": <1-5>,
      "required_precautions": ["<precaution that MUST be followed>"],
      "already_mentioned_precautions": ["<precautions the creator already mentions/shows>"],
      "missing_precautions": ["<precautions NOT mentioned but needed>"],
      "matched_rules": [
        {
          "rule_text": "<the safety rule>",
          "severity": <1-5>,
          "category": "<category>",
          "relevance": "<why this rule applies to this step>"
        }
      ]
    }
  ],
  "safety_measures_in_video": ["<all safety measures mentioned across entire video>"],
  "recommended_additional_measures": ["<measures not in video but strongly recommended>"]
}"""


def _build_safety_user_message(
    steps: list[dict[str, Any]],
    rules_per_step: dict[int, list[dict[str, Any]]],
    safety_categories: list[str],
    video_title: str = "",
) -> str:
    parts = []
    if video_title:
        parts.append(f"VIDEO TITLE: {video_title}\n")
    parts.append(f"SAFETY CATEGORIES: {', '.join(safety_categories)}\n")
    parts.append("=" * 60)
    parts.append("EXTRACTED DIY STEPS:")
    parts.append("=" * 60)

    for step in steps:
        step_num = step.get("step_number", "?")
        parts.append(f"\n--- Step {step_num} ---")
        parts.append(f"Action: {step.get('action_summary', 'N/A')}")
        parts.append(f"Instruction: {step.get('step_text', 'N/A')}")
        excerpt = step.get("transcript_excerpt", "")
        if excerpt:
            parts.append(f"Transcript: \"{excerpt}\"")
        matched = rules_per_step.get(step_num, [])
        if matched:
            parts.append(f"\nMATCHED SAFETY RULES ({len(matched)} rules):")
            for i, rule in enumerate(matched, 1):
                severity = rule.get("validated_severity") or rule.get("suggested_severity", 3)
                cats = rule.get("categories", [])
                sim = rule.get("similarity", 0)
                parts.append(
                    f"  {i}. [{severity}/5] [{', '.join(cats)}] "
                    f"(similarity: {sim:.0%}) {rule.get('actionable_rule', '')}"
                )
        else:
            parts.append("\nMATCHED SAFETY RULES: None found")

    return "\n".join(parts)


def _clean_json_response(text: str) -> str:
    trimmed = text.strip()
    if trimmed.startswith("```"):
        newline_pos = trimmed.find("\n")
        after_fence = trimmed[newline_pos + 1:] if newline_pos >= 0 else trimmed.lstrip("`")
        stripped = after_fence.rstrip("`").strip()
    else:
        stripped = trimmed
    obj_start = stripped.find("{")
    obj_end = stripped.rfind("}")
    if obj_start >= 0 and obj_end >= 0 and obj_end > obj_start:
        return stripped[obj_start: obj_end + 1]
    return stripped


async def analyze_safety(
    steps: list[dict[str, Any]],
    rules_per_step: dict[int, list[dict[str, Any]]],
    safety_categories: list[str],
    video_title: str = "",
    api_key: str | None = None,
    model: str = "qwen/qwen3-32b",
) -> dict[str, Any]:
    key = api_key or os.getenv("GROQ_API_KEY", "")
    if not key:
        raise RuntimeError("GROQ_API_KEY not set")

    user_message = _build_safety_user_message(steps, rules_per_step, safety_categories, video_title)
    is_qwen = "qwen" in model.lower()
    system_content = SAFETY_ANALYSIS_PROMPT + "\n\n/no_think" if is_qwen else SAFETY_ANALYSIS_PROMPT

    request_body: dict[str, Any] = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_content},
            {"role": "user", "content": user_message},
        ],
        "temperature": 0.1,
        "max_tokens": 8192,
        "stream": False,
        "seed": 42,
    }
    headers = {
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient() as client:
        resp = await client.post(GROQ_API_URL, json=request_body, headers=headers, timeout=120.0)
        if resp.status_code != 200:
            body = resp.text[:500]
            if resp.status_code == 401:
                raise Exception("Invalid Groq API key.")
            elif resp.status_code == 429:
                raise Exception("Groq rate limit exceeded. Wait a moment and try again.")
            elif resp.status_code == 503:
                raise Exception("Groq service temporarily unavailable.")
            else:
                raise Exception(f"Groq API error (HTTP {resp.status_code}): {body}")
        data = resp.json()
        raw_content = data["choices"][0]["message"]["content"] or ""

    if not raw_content.strip():
        raise Exception("Groq returned empty safety analysis response.")

    report = json.loads(_clean_json_response(raw_content))
    if not isinstance(report, dict):
        raise Exception(f"Expected JSON object, got: {type(report).__name__}")

    report.setdefault("verdict", "UNSAFE")
    report.setdefault("overall_risk_score", 3.0)
    report.setdefault("parent_monitoring_required", True)
    report.setdefault("parent_monitoring_reason", "")
    report.setdefault("summary", "")
    report.setdefault("critical_concerns", [])
    report.setdefault("step_safety_analysis", [])
    report.setdefault("safety_measures_in_video", [])
    report.setdefault("recommended_additional_measures", [])

    logger.info(
        "Safety analysis complete: verdict=%s, risk=%.1f, steps=%d",
        report["verdict"], report["overall_risk_score"],
        len(report["step_safety_analysis"]),
    )
    return report


# ---------------------------------------------------------------------------
# FastAPI routers
# ---------------------------------------------------------------------------

router = APIRouter(prefix="/api")
ws_router = APIRouter()
_extract_pool = ThreadPoolExecutor(max_workers=2, thread_name_prefix="extract")
_cache = AnalysisCache()

MAX_TRANSCRIPT_LENGTH = 100_000

ANALYSIS_MODELS = [
    {"key": "qwen", "model_id": "qwen/qwen3-32b", "label": "Qwen3 32B"},
    {"key": "gpt_oss", "model_id": "openai/gpt-oss-20b", "label": "GPT-OSS 20B"},
]


def _build_model_comparison(reports: dict[str, dict]) -> dict:
    comparison = {"models": [], "aspects": []}
    model_keys = []
    for m in ANALYSIS_MODELS:
        key = m["key"]
        if key in reports and reports[key]:
            model_keys.append(key)
            comparison["models"].append({"key": key, "label": m["label"]})

    if len(model_keys) < 2:
        return comparison

    def _aspect(name: str, extractor):
        values = {}
        for k in model_keys:
            try:
                values[k] = extractor(reports[k])
            except Exception:
                values[k] = "N/A"
        unique = set(str(v) for v in values.values() if v != "N/A")
        return {"aspect": name, "values": values, "agreement": len(unique) <= 1}

    comparison["aspects"] = [
        _aspect("Verdict", lambda r: r.get("verdict", "N/A")),
        _aspect("Overall Risk Score", lambda r: round(r.get("overall_risk_score", 0), 1)),
        _aspect("Parent Monitoring Required", lambda r: "Yes" if r.get("parent_monitoring_required") else "No"),
        _aspect("Critical Concerns Count", lambda r: len(r.get("critical_concerns", []))),
        _aspect("Total Missing Precautions", lambda r: sum(
            len(s.get("missing_precautions", [])) for s in r.get("step_safety_analysis", [])
        )),
        _aspect("Average Step Risk Level", lambda r: round(
            sum(s.get("risk_level", 0) for s in r.get("step_safety_analysis", []))
            / max(len(r.get("step_safety_analysis", [])), 1), 1
        )),
        _aspect("High-Risk Steps (>=4)", lambda r: sum(
            1 for s in r.get("step_safety_analysis", []) if s.get("risk_level", 0) >= 4
        )),
        _aspect("Total Matched Rules", lambda r: sum(
            len(s.get("matched_rules", [])) for s in r.get("step_safety_analysis", [])
        )),
        _aspect("Safety Measures Identified", lambda r: len(r.get("safety_measures_in_video", []))),
        _aspect("Recommended Additions", lambda r: len(r.get("recommended_additional_measures", []))),
        _aspect("Steps Analyzed", lambda r: len(r.get("step_safety_analysis", []))),
    ]
    return comparison


@router.get("/health")
async def health():
    key = get_api_key()
    db = get_database_url()
    return {
        "status": "ok",
        "api_key_configured": bool(key),
        "database_configured": bool(db),
        "model": get_model(),
    }


@router.get("/analyze")
async def analyze_diy(video_id: str = Query(...)):
    """SSE stream: transcript → LLM extraction → embeddings → rule match → safety report."""

    async def event_generator():
        try:
            api_key = get_api_key()
            if not api_key:
                yield {"event": "message", "data": json.dumps({"type": "error", "message": "GROQ_API_KEY not configured on server. Check .env file."})}
                return

            model = get_model()

            # Check cache
            cached = _cache.get(video_id)
            if cached:
                cached_data = json.loads(cached)
                async with httpx.AsyncClient() as client:
                    try:
                        meta = await fetch_metadata(client, video_id)
                        yield {"event": "message", "data": json.dumps({"type": "metadata", "title": meta.title, "author": meta.author})}
                    except Exception:
                        pass

                if cached_data.get("is_diy") is False:
                    yield {"event": "message", "data": json.dumps({"type": "not_diy", "message": "This video is not a DIY tutorial."})}
                    yield {"event": "message", "data": json.dumps({"type": "done"})}
                    return

                if "steps_json" in cached_data:
                    yield {"event": "message", "data": json.dumps({"type": "steps_complete", "steps_json": cached_data["steps_json"], "is_diy": True, "safety_categories": cached_data.get("safety_categories", [])})}

                cached_all = cached_data.get("all_reports_json", {})
                if cached_all:
                    for m in ANALYSIS_MODELS:
                        rpt = cached_all.get(m["key"])
                        if rpt:
                            yield {"event": "message", "data": json.dumps({"type": "safety_report", "model_key": m["key"], "model_label": m["label"], "report_json": rpt})}
                elif "report_json" in cached_data:
                    yield {"event": "message", "data": json.dumps({"type": "safety_report", "model_key": ANALYSIS_MODELS[0]["key"], "model_label": ANALYSIS_MODELS[0]["label"], "report_json": cached_data["report_json"]})}

                if "comparison_json" in cached_data:
                    yield {"event": "message", "data": json.dumps({"type": "model_comparison", "comparison_json": cached_data["comparison_json"]})}

                yield {"event": "message", "data": json.dumps({"type": "done"})}
                return

            # 1. Fetch transcript + metadata
            async with httpx.AsyncClient() as client:
                yield {"event": "message", "data": json.dumps({"type": "status", "message": "Fetching video transcript..."})}

                transcript_result, metadata_result = await asyncio.gather(
                    asyncio.create_task(fetch_transcript(client, video_id)),
                    asyncio.create_task(fetch_metadata(client, video_id)),
                    return_exceptions=True,
                )

                if isinstance(transcript_result, Exception):
                    yield {"event": "message", "data": json.dumps({"type": "error", "message": str(transcript_result)})}
                    return

                video_title = ""
                if not isinstance(metadata_result, Exception):
                    video_title = metadata_result.title
                    yield {"event": "message", "data": json.dumps({"type": "metadata", "title": metadata_result.title, "author": metadata_result.author})}

                if len(transcript_result.text) > MAX_TRANSCRIPT_LENGTH:
                    yield {"event": "message", "data": json.dumps({"type": "error", "message": f"Transcript is too long ({len(transcript_result.text) // 1000}k chars). Maximum is {MAX_TRANSCRIPT_LENGTH // 1000}k characters."})}
                    return

                # 2. Extract DIY steps via Groq (streaming)
                yield {"event": "message", "data": json.dumps({"type": "status", "message": "Extracting DIY steps from transcript..."})}

                steps_json = ""
                async for event in extract_steps_stream(client, api_key, model, transcript_result.text):
                    if event["type"] == "steps_delta":
                        yield {"event": "message", "data": json.dumps(event)}
                    elif event["type"] == "steps_complete":
                        steps_json = event["steps_json"]

                if not steps_json:
                    yield {"event": "message", "data": json.dumps({"type": "error", "message": "Failed to extract steps from transcript."})}
                    return

                # 3. Check is_diy
                parsed_extraction = json.loads(steps_json)
                is_diy = parsed_extraction.get("is_diy", True)
                safety_categories = parsed_extraction.get("safety_categories", ["general_safety"])
                steps_list = parsed_extraction.get("steps", []) if isinstance(parsed_extraction, dict) else parsed_extraction

                yield {"event": "message", "data": json.dumps({"type": "steps_complete", "steps_json": steps_json, "is_diy": is_diy, "safety_categories": safety_categories})}

                if not is_diy or not steps_list:
                    _cache.set(video_id, json.dumps({"is_diy": False}))
                    yield {"event": "message", "data": json.dumps({"type": "not_diy", "message": "This video is not a DIY tutorial. No safety analysis needed."})}
                    yield {"event": "message", "data": json.dumps({"type": "done"})}
                    return

                # 4. Embed steps
                yield {"event": "message", "data": json.dumps({"type": "status", "message": f"Generating embeddings for {len(steps_list)} steps..."})}
                embed_service = await asyncio.to_thread(EmbeddingService.get_instance)
                step_embeddings = await asyncio.to_thread(embed_service.embed_steps, steps_list)

                # 5. Match against safety rules via pgvector
                yield {"event": "message", "data": json.dumps({"type": "status", "message": "Matching steps against safety rules database..."})}

                rules_per_step: dict[int, list] = {}
                for step, embedding in zip(steps_list, step_embeddings):
                    step_num = step.get("step_number", 0)
                    try:
                        matched = await asyncio.to_thread(embed_service.find_rules_for_step, step, embedding, safety_categories)
                        rules_per_step[step_num] = matched
                    except Exception as e:
                        rules_per_step[step_num] = []
                        yield {"event": "message", "data": json.dumps({"type": "status", "message": f"Rule matching for step {step_num} skipped: {e}"})}

                total_matched = sum(len(v) for v in rules_per_step.values())
                yield {"event": "message", "data": json.dumps({"type": "status", "message": f"Found {total_matched} matching rules across {len(steps_list)} steps. Running safety assessment..."})}

                # 6. Multi-model safety assessment (parallel)
                yield {"event": "message", "data": json.dumps({"type": "status", "message": f"Running safety assessment across {len(ANALYSIS_MODELS)} models..."})}

                async def _run_model(m: dict) -> tuple[str, dict | None, str | None]:
                    try:
                        r = await analyze_safety(
                            steps=steps_list, rules_per_step=rules_per_step,
                            safety_categories=safety_categories, video_title=video_title,
                            api_key=api_key, model=m["model_id"],
                        )
                        return (m["key"], r, None)
                    except Exception as exc:
                        return (m["key"], None, str(exc))

                model_results = await asyncio.gather(*[_run_model(m) for m in ANALYSIS_MODELS])

                all_reports: dict[str, dict] = {}
                all_reports_json: dict[str, str] = {}
                primary_report_json = "{}"

                for key, rpt, err in model_results:
                    if rpt:
                        all_reports[key] = rpt
                        rpt_json = json.dumps(rpt)
                        all_reports_json[key] = rpt_json
                        label = next((m["label"] for m in ANALYSIS_MODELS if m["key"] == key), key)
                        yield {"event": "message", "data": json.dumps({"type": "safety_report", "model_key": key, "model_label": label, "report_json": rpt_json})}
                        if key == ANALYSIS_MODELS[0]["key"]:
                            primary_report_json = rpt_json
                    else:
                        label = next((m["label"] for m in ANALYSIS_MODELS if m["key"] == key), key)
                        yield {"event": "message", "data": json.dumps({"type": "status", "message": f"Safety assessment error for {label}: {err}"})}

                # 6b. Comparison table
                comparison = _build_model_comparison(all_reports)
                comparison_json = json.dumps(comparison)
                yield {"event": "message", "data": json.dumps({"type": "model_comparison", "comparison_json": comparison_json})}

                # 7. Cache result
                _cache.set(video_id, json.dumps({
                    "is_diy": True,
                    "steps_json": steps_json,
                    "safety_categories": safety_categories,
                    "report_json": primary_report_json,
                    "all_reports_json": {k: v for k, v in all_reports_json.items()},
                    "comparison_json": comparison_json,
                }))

                yield {"event": "message", "data": json.dumps({"type": "done"})}

        except Exception as e:
            yield {"event": "message", "data": json.dumps({"type": "error", "message": str(e)})}

    return EventSourceResponse(event_generator())


@router.get("/rules")
async def get_rules(
    category: Optional[str] = Query(None),
    severity: Optional[int] = Query(None),
    document: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    run_id: Optional[int] = Query(None),
    page: int = Query(1),
    per_page: int = Query(50),
):
    try:
        if run_id is not None:
            return await asyncio.to_thread(fetch_rules_by_run, run_id=run_id, page=page, per_page=per_page)
        return await asyncio.to_thread(fetch_rules_from_db, category=category, severity=severity, document=document, search=search, page=page, per_page=per_page)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/filter_options")
async def get_filter_options():
    try:
        return await asyncio.to_thread(fetch_filter_options_from_db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rules_by_document")
async def get_rules_by_document_endpoint():
    try:
        return await asyncio.to_thread(fetch_rules_by_document)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/extraction_runs")
async def get_extraction_runs():
    try:
        return await asyncio.to_thread(fetch_extraction_runs)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/evaluation_results")
async def get_evaluation_results(run_id: Optional[int] = Query(default=None)):
    """Fetch per-file evaluation results, optionally filtered by run_id."""
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if run_id is not None:
                cur.execute(
                    "SELECT * FROM evaluation_results WHERE run_id = %s ORDER BY id",
                    (run_id,),
                )
            else:
                cur.execute("SELECT * FROM evaluation_results ORDER BY id DESC LIMIT 200")
            rows = cur.fetchall()

        results = []
        for row in rows:
            r = dict(row)
            if r.get("created_at"):
                r["created_at"] = r["created_at"].isoformat()
            if r.get("failed_rules") and isinstance(r["failed_rules"], str):
                r["failed_rules"] = json.loads(r["failed_rules"])
            results.append(r)

        return {"results": results}
    finally:
        conn.close()


@router.post("/system_eval/run")
async def run_system_eval(request: Request, sample_size: int = Query(50, ge=1, le=500)):
    try:
        payload = {}
        try:
            payload = await request.json()
        except Exception:
            payload = {}

        urls = payload.get("youtube_urls") or []
        if isinstance(urls, str):
            urls = _extract_urls_from_text_blob(urls)
        elif isinstance(urls, list):
            normed = []
            for u in urls:
                if isinstance(u, str):
                    n = _normalize_youtube_url(u)
                    if n:
                        normed.append(n)
            urls = list(dict.fromkeys(normed))
        else:
            urls = []

        if not urls and payload.get("use_pool", True):
            urls = _get_url_pool()

        random_count = payload.get("random_count")
        random_min = payload.get("random_min")
        random_max = payload.get("random_max")
        if urls and random_min is not None and random_max is not None:
            lo = max(1, int(random_min))
            hi = max(lo, int(random_max))
            hi = min(hi, len(urls))
            random_count = random.randint(lo, hi)

        return await asyncio.to_thread(_evaluate_system, sample_size, urls, random_count)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _extract_urls_from_upload(name: str, data: bytes) -> list[str]:
    lower = (name or "").lower()

    if lower.endswith(".txt"):
        return _extract_urls_from_text_blob(data.decode("utf-8", errors="ignore"))

    if lower.endswith(".csv"):
        text = data.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        urls: list[str] = []
        for row in reader:
            if not row:
                continue
            normalized = {str(k).strip().lower(): v for k, v in row.items()}
            candidate = normalized.get("youtube_url") or normalized.get("url")
            if candidate:
                urls.extend(_extract_urls_from_text_blob(str(candidate)))
        return list(dict.fromkeys(urls))

    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except Exception:
            return []
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        headers = []
        urls: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = ["" if v is None else str(v).strip() for v in row]
            if i == 1:
                headers = [h.lower() for h in vals]
                continue
            if not headers:
                continue
            row_map = {headers[idx]: vals[idx] for idx in range(min(len(headers), len(vals)))}
            candidate = row_map.get("youtube_url") or row_map.get("url")
            if candidate:
                urls.extend(_extract_urls_from_text_blob(candidate))
        return list(dict.fromkeys(urls))

    if lower.endswith(".pdf"):
        text_chunks: list[str] = []
        try:
            try:
                import pymupdf as fitz
            except Exception:
                import fitz  # type: ignore
            doc = fitz.open(stream=data, filetype="pdf")
            for page in doc:
                txt = page.get_text() or ""
                text_chunks.append(txt)
                for link in page.get_links() or []:
                    uri = link.get("uri") if isinstance(link, dict) else None
                    if uri:
                        text_chunks.append(str(uri))
            doc.close()
        except Exception:
            return []
        return _extract_urls_from_text_blob("\n".join(text_chunks))

    return _extract_urls_from_text_blob(data.decode("utf-8", errors="ignore"))


@router.post("/system_eval/collect_urls")
async def collect_system_eval_urls(files: List[UploadFile] = File(default=[]), pasted_urls: str = Form(default="")):
    try:
        all_urls: list[str] = []
        source_files: list[str] = []

        if pasted_urls:
            all_urls.extend(_extract_urls_from_text_blob(pasted_urls))

        for f in files or []:
            content = await f.read()
            source_files.append(f.filename or "")
            all_urls.extend(_extract_urls_from_upload(f.filename or "", content))

        all_urls = list(dict.fromkeys([u for u in all_urls if u]))
        inserted, total_pool = _insert_urls_into_bucket(all_urls, source_type="eval_input", source_file=", ".join([s for s in source_files if s]) or None)

        return {
            "added_urls": all_urls,
            "added_count": inserted,
            "total_urls_in_pool": total_pool,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/system_eval/url_pool")
async def get_system_eval_url_pool():
    try:
        urls = await asyncio.to_thread(_get_url_pool)
        return {"total_urls": len(urls), "urls": urls}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/system_eval/latest")
async def get_latest_system_eval():
    try:
        latest = await asyncio.to_thread(_fetch_latest_system_evaluation)
        if not latest:
            return {"result": None}
        return {"result": latest}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/extract_rules")
async def extract_rules_endpoint(files: List[UploadFile] = File(default=[])):
    if not files:
        raise HTTPException(status_code=400, detail="No files provided.")

    results = []
    errors = []

    for file in files:
        if not file.filename or not file.filename.lower().endswith(".pdf"):
            errors.append({"file": file.filename or "unknown", "error": "Not a PDF file"})
            continue

        tmp = Path(tempfile.gettempdir()) / file.filename
        content = await file.read()
        tmp.write_bytes(content)

        try:
            result = await asyncio.to_thread(extract_rules_v2, str(tmp), file.filename)
            results.append({"file": file.filename, "run_id": result["run_id"], "extraction": result["extraction"], "evaluation_results": result["evaluation_results"]})
        except Exception as e:
            errors.append({"file": file.filename, "error": str(e)})
        finally:
            try:
                tmp.unlink()
            except OSError:
                pass

    return {"results": results, "errors": errors, "total_files": len(files), "successful": len(results), "failed": len(errors)}


@router.post("/run_evaluation/{run_id}")
async def trigger_evaluation(run_id: int):
    try:
        conn = get_db_connection()
        try:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("SELECT id, source_documents, json_source_file FROM extraction_runs WHERE id = %s", (run_id,))
                run = cur.fetchone()
                if not run:
                    raise HTTPException(status_code=404, detail=f"Run #{run_id} not found")

            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """SELECT rule_id, original_text, actionable_rule, materials,
                              suggested_severity, validated_severity, categories,
                              source_document, page_number, section_heading
                       FROM safety_rules WHERE run_id = %s""",
                    (run_id,),
                )
                rules = [dict(r) for r in cur.fetchall()]
        finally:
            conn.close()

        if not rules:
            return {"run_id": run_id, "error": "No rules found for this run"}

        file_name = run.get("json_source_file") or "unknown"
        evaluation = await asyncio.to_thread(run_structure_evaluation, {"rules": rules})
        await asyncio.to_thread(save_evaluation_results, run_id, evaluation, file_name)
        return {"run_id": run_id, "evaluation_results": evaluation}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/scans")
async def save_scan(request: Request):
    body = await request.json()
    video_id = body.get("video_id", "")
    video_url = body.get("video_url", "")
    title = body.get("title", "")
    channel = body.get("channel", "")
    verdict = body.get("verdict", "")
    risk_score = body.get("risk_score")
    output_json = body.get("output_json", {})
    model_reports = output_json.get("modelReports")
    comparison_data = output_json.get("comparison")

    if not video_id or not title:
        raise HTTPException(status_code=400, detail="video_id and title are required")

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO completed_scans
                    (video_id, video_url, title, channel, verdict, risk_score, output_json,
                     model_reports, comparison_data)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, scan_timestamp
                """,
                (video_id, video_url, title, channel, verdict, risk_score,
                 psycopg2.extras.Json(output_json),
                 psycopg2.extras.Json(model_reports) if model_reports else None,
                 psycopg2.extras.Json(comparison_data) if comparison_data else None),
            )
            row = cur.fetchone()
            conn.commit()
            return {"id": row[0], "scan_timestamp": row[1].isoformat()}
    finally:
        conn.close()


@router.get("/scans")
async def list_scans():
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, video_id, video_url, title, channel, verdict,
                       risk_score, scan_timestamp
                FROM completed_scans
                ORDER BY scan_timestamp DESC
                LIMIT 200
                """
            )
            rows = cur.fetchall()
            return {
                "scans": [
                    {
                        "id": r[0], "video_id": r[1], "video_url": r[2],
                        "title": r[3], "channel": r[4], "verdict": r[5],
                        "risk_score": r[6],
                        "scan_timestamp": r[7].isoformat() if r[7] else None,
                    }
                    for r in rows
                ]
            }
    finally:
        conn.close()


@router.get("/scans/{scan_id}")
async def get_scan(scan_id: int):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, video_id, video_url, title, channel, verdict,
                       risk_score, scan_timestamp, output_json
                FROM completed_scans WHERE id = %s
                """,
                (scan_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Scan not found")
            return {
                "id": row[0], "video_id": row[1], "video_url": row[2],
                "title": row[3], "channel": row[4], "verdict": row[5],
                "risk_score": row[6],
                "scan_timestamp": row[7].isoformat() if row[7] else None,
                "output_json": row[8],
            }
    finally:
        conn.close()


@ws_router.websocket("/ws/extract")
async def ws_extract(ws: WebSocket):
    """WebSocket extraction with real-time progress."""
    await ws.accept()
    try:
        raw = await ws.receive_text()
        msg = json.loads(raw)
        files_data = msg.get("files", [])

        if not files_data:
            await ws.send_json({"step": "error", "status": "No files provided"})
            await ws.close()
            return

        loop = asyncio.get_event_loop()
        all_results = []

        for file_info in files_data:
            file_name = file_info.get("name", "unknown.pdf")
            file_b64 = file_info.get("data", "")

            if not file_b64:
                await ws.send_json({"step": "error", "status": f"No data for file: {file_name}"})
                continue

            file_bytes = base64.b64decode(file_b64)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf", prefix="extract_")
            tmp.write(file_bytes)
            tmp.close()
            tmp_path = tmp.name

            import queue
            progress_queue: queue.Queue = queue.Queue()

            def progress_callback(step: str, detail: dict):
                progress_queue.put({"step": step, **detail})

            def run_extraction():
                try:
                    return extract_rules_with_progress(tmp_path, file_name, progress_callback)
                except Exception as e:
                    progress_queue.put({"step": "error", "status": f"Extraction failed: {str(e)}"})
                    return None

            future = loop.run_in_executor(_extract_pool, run_extraction)

            while not future.done():
                try:
                    event = progress_queue.get_nowait()
                    await ws.send_json(event)
                except queue.Empty:
                    pass
                await asyncio.sleep(0.1)

            while not progress_queue.empty():
                event = progress_queue.get_nowait()
                await ws.send_json(event)

            result = future.result()
            if result:
                all_results.append(result)

            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        await ws.send_json({
            "step": "done",
            "status": f"All files processed ({len(all_results)} succeeded)",
            "results": [
                {
                    "run_id": r.get("run_id"),
                    "rule_count": r.get("extraction", {}).get("rule_count", 0),
                    "accuracy": r.get("evaluation_results", {}).get("overall_accuracy"),
                }
                for r in all_results
            ],
        })

    except WebSocketDisconnect:
        pass
