from __future__ import annotations

import argparse
import atexit
import csv
import io
import json
import math
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, urlparse, urlsplit, urlunsplit

import httpx
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv


ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / "backend" / ".env")

YOUTUBE_URL_RE = re.compile(r'https?://(?:www\.)?(?:youtube\.com|youtu\.be)/[^\s,;\]\)>"\']+', re.IGNORECASE)


@dataclass
class VideoEvalResult:
    url_id: str
    video_id: str
    video_url: str
    title: str
    channel: str
    scan_id: int | None
    ground_truth_label: str | None
    predicted_label: str | None
    label_match: bool | None
    steps_evaluated: int
    total_precautions: int
    supported_precautions: int
    step_accuracy: float
    step_precision: float
    step_recall: float
    step_f1_score: float
    label_tp: int
    label_tn: int
    label_fp: int
    label_fn: int
    mrr: float
    faithfulness: float
    spearman: float
    risk_score: float
    average_severity_score: float


@dataclass
class PoolEntry:
    id: str
    youtube_url: str
    video_id: str
    title: str | None
    categories: str | None
    safety_label: str | None


def get_db_connection():
    url = (os.getenv("DATABASE_URL") or os.getenv("SUPABASE_URL", "")).strip()
    if not url:
        raise RuntimeError("DATABASE_URL or SUPABASE_URL is not configured")
    # Handle raw passwords that may contain spaces/special characters in .env.
    try:
        parts = urlsplit(url)
        netloc = parts.netloc
        if "@" in netloc:
            auth, host = netloc.rsplit("@", 1)
            if ":" in auth:
                user, pwd = auth.split(":", 1)
                auth = f"{user}:{quote(pwd, safe='')}"
                netloc = f"{auth}@{host}"
                url = urlunsplit((parts.scheme, netloc, parts.path, parts.query, parts.fragment))
    except Exception:
        pass

    if "supabase.com" in url and "sslmode=" not in url:
        joiner = "&" if "?" in url else "?"
        url = f"{url}{joiner}sslmode=require"

    # Supabase pooler connections often require port 6543 instead of 5432.
    candidate_urls = [url]
    if "pooler.supabase.com" in url and ":5432" in url:
        candidate_urls.append(url.replace(":5432", ":6543", 1))

    last_exc: Exception | None = None
    for candidate in candidate_urls:
        try:
            return psycopg2.connect(candidate, connect_timeout=10)
        except Exception as exc:
            last_exc = exc

    if last_exc:
        raise last_exc
    raise RuntimeError("Failed to establish database connection")


def ensure_system_eval_schema() -> None:
    """Ensure required Supabase tables/columns exist for system eval flow."""
    conn = get_db_connection()
    migrations = [
        "CREATE EXTENSION IF NOT EXISTS pgcrypto;",
        """
        CREATE TABLE IF NOT EXISTS youtube_urls (
            id              UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            youtube_url     TEXT NOT NULL UNIQUE,
            video_id        TEXT,
            title           TEXT,
            categories      TEXT,
            safety_label    TEXT,
            created_at      TIMESTAMPTZ DEFAULT now()
        );
        """,
        "CREATE INDEX IF NOT EXISTS idx_youtube_urls_video_id ON youtube_urls (video_id);",
        "ALTER TABLE youtube_urls ADD COLUMN IF NOT EXISTS youtube_url TEXT;",
        "ALTER TABLE youtube_urls ADD COLUMN IF NOT EXISTS safety_label TEXT;",
        "ALTER TABLE youtube_urls DROP COLUMN IF EXISTS source_type;",
        "ALTER TABLE youtube_urls DROP COLUMN IF EXISTS source_file;",
        "ALTER TABLE youtube_urls DROP COLUMN IF EXISTS ground_truth_binary;",
        "ALTER TABLE youtube_urls DROP COLUMN IF EXISTS last_used_at;",
        """
        DO $$
        BEGIN
            IF EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'youtube_urls' AND column_name = 'url'
            ) THEN
                UPDATE youtube_urls
                SET youtube_url = url
                WHERE youtube_url IS NULL AND url IS NOT NULL;
            END IF;
        END $$;
        """,
        """
        DO $$
        BEGIN
            IF EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'youtube_urls' AND column_name = 'ground_truth_label'
            ) THEN
                UPDATE youtube_urls
                SET safety_label = ground_truth_label
                WHERE safety_label IS NULL AND ground_truth_label IS NOT NULL;
            END IF;
        END $$;
        """,
        "ALTER TABLE youtube_urls DROP COLUMN IF EXISTS url;",
        "ALTER TABLE youtube_urls DROP COLUMN IF EXISTS ground_truth_label;",
        """
        CREATE TABLE IF NOT EXISTS system_eval (
            id                      UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            evaluated_at            TIMESTAMPTZ DEFAULT now(),
            batch_start_index       INTEGER,
            batch_end_index         INTEGER,
            bucket_source_url       TEXT,
            accuracy                REAL,
            precision               REAL,
            recall                  REAL,
            f1_score                REAL,
            avg_mrr                 REAL,
            avg_faithfulness        REAL,
            avg_spearman            REAL
        );
        """,
        "CREATE INDEX IF NOT EXISTS idx_system_eval_evaluated_at ON system_eval (evaluated_at DESC);",
        """
        DO $$
        BEGIN
            IF EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_name = 'system_eval_video_results'
            ) AND NOT EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_name = 'individual_video_results'
            ) THEN
                ALTER TABLE system_eval_video_results RENAME TO individual_video_results;
            END IF;
        END $$;
        """,
        """
        CREATE TABLE IF NOT EXISTS individual_video_results (
            id                      UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            eval_id                 UUID NOT NULL REFERENCES system_eval(id) ON DELETE CASCADE,
            url_id                  UUID NOT NULL REFERENCES youtube_urls(id) ON DELETE CASCADE,
            predicted_label         TEXT,
            risk_score              REAL,
            total_steps             INTEGER,
            average_severity_score  REAL,
            spearman_correlation    REAL,
            faithfulness_score      REAL,
            created_at              TIMESTAMPTZ DEFAULT now()
        );
        """,
        "CREATE INDEX IF NOT EXISTS idx_individual_video_eval_id ON individual_video_results (eval_id);",
        "CREATE INDEX IF NOT EXISTS idx_individual_video_url_id ON individual_video_results (url_id);",
        "ALTER TABLE individual_video_results ADD COLUMN IF NOT EXISTS url_id UUID;",
        "ALTER TABLE individual_video_results ADD COLUMN IF NOT EXISTS predicted_label TEXT;",
        "ALTER TABLE individual_video_results ADD COLUMN IF NOT EXISTS risk_score REAL;",
        "ALTER TABLE individual_video_results ADD COLUMN IF NOT EXISTS total_steps INTEGER;",
        "ALTER TABLE individual_video_results ADD COLUMN IF NOT EXISTS average_severity_score REAL;",
        "ALTER TABLE individual_video_results ADD COLUMN IF NOT EXISTS spearman_correlation REAL;",
        "ALTER TABLE individual_video_results ADD COLUMN IF NOT EXISTS faithfulness_score REAL;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS video_id;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS video_url;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS ground_truth_label;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS label_match;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS scan_id;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS steps_evaluated;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS total_precautions;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS supported_precautions;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS true_positive;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS true_negative;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS false_positive;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS false_negative;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS accuracy;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS precision;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS recall;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS f1_score;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS mrr;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS faithfulness;",
        "ALTER TABLE individual_video_results DROP COLUMN IF EXISTS spearman;",
        """
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1
                FROM information_schema.table_constraints
                WHERE table_name = 'individual_video_results'
                  AND constraint_name = 'fk_individual_video_url_id'
            ) THEN
                ALTER TABLE individual_video_results
                ADD CONSTRAINT fk_individual_video_url_id
                FOREIGN KEY (url_id) REFERENCES youtube_urls(id) ON DELETE CASCADE;
            END IF;
        END $$;
        """,
    ]
    try:
        with conn.cursor() as cur:
            for sql in migrations:
                cur.execute(sql)
        conn.commit()
    finally:
        conn.close()


def verify_supabase_state() -> dict[str, Any]:
    """Return a compact verification summary from Supabase tables."""
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT COUNT(*) AS c FROM youtube_urls")
            pool_count = int((cur.fetchone() or {}).get("c") or 0)

            cur.execute("SELECT COUNT(*) AS c FROM youtube_urls WHERE safety_label IS NOT NULL")
            labeled_pool_count = int((cur.fetchone() or {}).get("c") or 0)

            cur.execute("SELECT COUNT(*) AS c FROM system_eval")
            eval_count = int((cur.fetchone() or {}).get("c") or 0)

            cur.execute("SELECT COUNT(*) AS c FROM individual_video_results")
            video_eval_count = int((cur.fetchone() or {}).get("c") or 0)

            cur.execute(
                """
                  SELECT id, evaluated_at, batch_start_index, batch_end_index,
                      accuracy, precision, recall, f1_score,
                      avg_mrr, avg_faithfulness, avg_spearman
                FROM system_eval
                ORDER BY evaluated_at DESC
                LIMIT 1
                """
            )
            latest_eval = dict(cur.fetchone() or {})
            if latest_eval.get("evaluated_at"):
                latest_eval["evaluated_at"] = latest_eval["evaluated_at"].isoformat()

        return {
            "youtube_urls_count": pool_count,
            "youtube_urls_labeled_count": labeled_pool_count,
            "system_eval_count": eval_count,
            "individual_video_results_count": video_eval_count,
            "latest_system_eval": latest_eval or None,
        }
    finally:
        conn.close()


def extract_video_id_from_url(value: str) -> str | None:
    raw = (value or "").strip()
    if not raw:
        return None
    if re.fullmatch(r"[A-Za-z0-9_-]{11}", raw):
        return raw
    try:
        p = urlparse(raw)
        host = p.netloc.lower()
        path = p.path.strip("/")
        if "youtu.be" in host:
            vid = path.split("/")[0]
            return vid if re.fullmatch(r"[A-Za-z0-9_-]{11}", vid or "") else None
        if "youtube.com" in host:
            qs = parse_qs(p.query)
            vid = (qs.get("v") or [None])[0]
            if vid and re.fullmatch(r"[A-Za-z0-9_-]{11}", vid):
                return vid
            if path.startswith("shorts/"):
                chunks = path.split("/")
                vid = chunks[1] if len(chunks) > 1 else None
                return vid if re.fullmatch(r"[A-Za-z0-9_-]{11}", vid or "") else None
    except Exception:
        return None
    return None


def normalize_youtube_url(value: str) -> str | None:
    vid = extract_video_id_from_url(value)
    if not vid:
        return None
    return f"https://www.youtube.com/watch?v={vid}"


def _normalize_header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (value or "").lower())


def _pick_header_value(row_map: dict[str, str], exact_keys: list[str], contains_any: list[str]) -> str | None:
    for key in exact_keys:
        val = row_map.get(key)
        if val:
            return val
    for k, v in row_map.items():
        if not v:
            continue
        if all(token in k for token in contains_any):
            return v
    return None


def _sanitize_db_text(value: str | None) -> str | None:
    """Normalize text for PostgreSQL inserts and strip invalid/null characters."""
    if value is None:
        return None
    text = str(value).replace("\x00", "")
    text = text.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore").strip()
    return text or None


def _normalize_ground_truth_label(value: str | None) -> str | None:
    raw = (value or "").strip().upper()
    if not raw:
        return None
    compact = re.sub(r"[^A-Z]", "", raw)
    if "UNSAFE" in compact:
        return "UNSAFE"
    if "PSUA" in compact or "ADULTSUPERVISION" in compact or "SUPERVISION" in compact:
        return "PSUA"
    if "SAFE" in compact:
        return "SAFE"
    return None


def _ground_truth_to_binary(label: str | None) -> int | None:
    if label == "SAFE":
        return 0
    if label in {"UNSAFE", "PSUA"}:
        return 1
    return None


def _extract_url_and_label_pairs_from_text(text: str) -> list[tuple[str, str | None]]:
    pairs: list[tuple[str, str | None]] = []
    if not text:
        return pairs

    label_re = r"(SAFE|UNSAFE|PSUA)"
    url_re = r"(https?://(?:www\.)?(?:youtube\.com|youtu\.be)/[^\s,;\]\)>\"']+)"

    for m in re.finditer(rf"{url_re}\s*[:|,;\t ]*{label_re}", text, flags=re.IGNORECASE):
        pairs.append((m.group(1), _normalize_ground_truth_label(m.group(2))))
    for m in re.finditer(rf"{label_re}\s*[:|,;\t ]*{url_re}", text, flags=re.IGNORECASE):
        pairs.append((m.group(2), _normalize_ground_truth_label(m.group(1))))

    if pairs:
        return pairs

    # Fallback for table-like PDF exports where URL and label are split across
    # separate cells/lines and lose row structure in plain extracted text.
    urls_in_order = YOUTUBE_URL_RE.findall(text)
    labels_in_order: list[str] = []
    for m in re.finditer(r"\b(SAFE|UNSAFE|PSUA)\b", text, flags=re.IGNORECASE):
        norm = _normalize_ground_truth_label(m.group(1))
        if norm:
            labels_in_order.append(norm)

    if urls_in_order and labels_in_order:
        # Pair by order only when enough labels are present to be useful.
        # This works well for spreadsheet/table PDFs exported to text.
        if len(labels_in_order) >= max(1, len(urls_in_order) // 2):
            for idx, u in enumerate(urls_in_order):
                label = labels_in_order[idx] if idx < len(labels_in_order) else None
                pairs.append((u, label))
            return pairs

    for u in YOUTUBE_URL_RE.findall(text):
        pairs.append((u, None))
    return pairs


def extract_urls_from_text_blob(text: str) -> list[str]:
    if not text:
        return []
    found = YOUTUBE_URL_RE.findall(text)
    tokens = re.split(r"[\s,;]+", text)
    for token in tokens:
        if re.fullmatch(r"[A-Za-z0-9_-]{11}", token or ""):
            found.append(token)

    deduped: list[str] = []
    seen = set()
    for item in found:
        n = normalize_youtube_url(item)
        if n and n not in seen:
            deduped.append(n)
            seen.add(n)
    return deduped


def _build_pool_entry(url: str, title: str | None = None, categories: str | None = None, label: str | None = None) -> PoolEntry | None:
    norm = normalize_youtube_url(_sanitize_db_text(url) or "")
    if not norm:
        return None
    video_id = extract_video_id_from_url(norm)
    if not video_id:
        return None
    gt = _normalize_ground_truth_label(label)
    return PoolEntry(
        id="",
        youtube_url=norm,
        video_id=video_id,
        title=_sanitize_db_text(title),
        categories=_sanitize_db_text(categories),
        safety_label=gt,
    )


def _dedup_entries(entries: list[PoolEntry]) -> list[PoolEntry]:
    dedup: dict[str, PoolEntry] = {}
    for e in entries:
        prev = dedup.get(e.youtube_url)
        if prev is None:
            dedup[e.youtube_url] = e
            continue
        # Prefer the version that carries a label.
        if (not prev.safety_label) and e.safety_label:
            dedup[e.youtube_url] = e
    return list(dedup.values())


def _extract_row_labels_from_pdf_text(text: str) -> list[str]:
    labels: list[str] = []
    if not text:
        return labels
    # Rows in this source typically start with VID_###. Grab the first label in each row chunk.
    parts = re.split(r"\bVID_\d+\b", text, flags=re.IGNORECASE)
    for chunk in parts[1:]:
        m = re.search(r"\b(SAFE|UNSAFE|PSUA)\b", chunk, flags=re.IGNORECASE)
        if not m:
            continue
        norm = _normalize_ground_truth_label(m.group(1))
        if norm:
            labels.append(norm)
    return labels


def extract_entries_from_pdf(path: Path) -> list[PoolEntry]:
    try:
        try:
            import pymupdf as fitz
        except Exception:
            import fitz  # type: ignore

        doc = fitz.open(str(path))
        text_chunks: list[str] = []
        link_uris: list[str] = []
        for page in doc:
            text_chunks.append(page.get_text() or "")
            for link in page.get_links() or []:
                if isinstance(link, dict) and link.get("uri"):
                    link_uris.append(str(link["uri"]))
        doc.close()

        text_blob = "\n".join(text_chunks)

        entries: list[PoolEntry] = []
        # Parse direct text first (keeps local URL/label context when available).
        for url, label in _extract_url_and_label_pairs_from_text(text_blob):
            entry = _build_pool_entry(url, label=label)
            if entry:
                entries.append(entry)

        # Fallback: pair ordered unique clickable URLs with row labels extracted from text.
        if link_uris:
            ordered_unique_urls: list[str] = []
            seen_video_ids: set[str] = set()
            for raw_url in link_uris:
                norm = normalize_youtube_url(raw_url)
                vid = extract_video_id_from_url(norm or "") if norm else None
                if not norm or not vid or vid in seen_video_ids:
                    continue
                seen_video_ids.add(vid)
                ordered_unique_urls.append(norm)

            row_labels = _extract_row_labels_from_pdf_text(text_blob)
            if ordered_unique_urls and row_labels:
                for idx, norm_url in enumerate(ordered_unique_urls):
                    label = row_labels[idx] if idx < len(row_labels) else None
                    entry = _build_pool_entry(norm_url, label=label)
                    if entry:
                        entries.append(entry)

        return _dedup_entries(entries)
    except Exception:
        return []


def extract_entries_from_csv(path: Path) -> list[PoolEntry]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    entries: list[PoolEntry] = []
    for row in reader:
        if not row:
            continue
        normalized = {_normalize_header_key(str(k)): ("" if v is None else str(v).strip()) for k, v in row.items()}
        candidate = _pick_header_value(
            normalized,
            exact_keys=["youtubeurl", "url", "titleurl"],
            contains_any=["url"],
        )
        title = _pick_header_value(
            normalized,
            exact_keys=["title"],
            contains_any=["title"],
        )
        categories = _pick_header_value(
            normalized,
            exact_keys=["categories", "category"],
            contains_any=["categor"],
        )
        label = _pick_header_value(
            normalized,
            exact_keys=["safetylabel", "safeunsafepsua", "label", "safety"],
            contains_any=["safe"],
        )
        if not label:
            label = _pick_header_value(
                normalized,
                exact_keys=[],
                contains_any=["label"],
            )
        if candidate:
            for u in extract_urls_from_text_blob(str(candidate)):
                entry = _build_pool_entry(u, title=title, categories=categories, label=label)
                if entry:
                    entries.append(entry)
    return _dedup_entries(entries)


def extract_entries_from_excel(path: Path) -> list[PoolEntry]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    entries: list[PoolEntry] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = ["" if v is None else str(v).strip() for v in row]
        if i == 1:
            headers = [_normalize_header_key(h) for h in vals]
            continue
        if not headers:
            continue
        row_map = {headers[idx]: vals[idx] for idx in range(min(len(headers), len(vals)))}
        candidate = _pick_header_value(
            row_map,
            exact_keys=["youtubeurl", "url", "titleurl"],
            contains_any=["url"],
        )
        title = _pick_header_value(
            row_map,
            exact_keys=["title"],
            contains_any=["title"],
        )
        categories = _pick_header_value(
            row_map,
            exact_keys=["categories", "category"],
            contains_any=["categor"],
        )
        label = _pick_header_value(
            row_map,
            exact_keys=["safetylabel", "safeunsafepsua", "label", "safety"],
            contains_any=["safe"],
        )
        if not label:
            label = _pick_header_value(
                row_map,
                exact_keys=[],
                contains_any=["label"],
            )
        if candidate:
            for u in extract_urls_from_text_blob(candidate):
                entry = _build_pool_entry(u, title=title, categories=categories, label=label)
                if entry:
                    entries.append(entry)
    return _dedup_entries(entries)


def extract_entries_from_file(path: Path) -> list[PoolEntry]:
    lower = path.name.lower()
    if lower.endswith(".pdf"):
        return extract_entries_from_pdf(path)
    if lower.endswith(".csv"):
        return extract_entries_from_csv(path)
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return extract_entries_from_excel(path)
    if lower.endswith(".txt"):
        txt = path.read_text(encoding="utf-8", errors="ignore")
        pairs = _extract_url_and_label_pairs_from_text(txt)
        items: list[PoolEntry] = []
        for u, label in pairs:
            entry = _build_pool_entry(u, label=label)
            if entry:
                items.append(entry)
        return _dedup_entries(items)
    return []


def collect_and_upsert_urls(urls_dir: Path) -> tuple[list[PoolEntry], int]:
    ensure_system_eval_schema()
    entries: list[PoolEntry] = []
    if urls_dir.exists():
        for p in sorted(urls_dir.iterdir()):
            if p.is_file() and not p.name.startswith("."):
                entries.extend(extract_entries_from_file(p))

    entries = _dedup_entries(entries)

    if not entries:
        return [], get_pool_size()

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            for e in entries:
                db_url = _sanitize_db_text(e.youtube_url)
                db_video_id = _sanitize_db_text(e.video_id)
                db_title = _sanitize_db_text(e.title)
                db_categories = _sanitize_db_text(e.categories)
                db_gt = _sanitize_db_text(e.safety_label)
                if not db_url:
                    continue
                cur.execute(
                    """
                    INSERT INTO youtube_urls (youtube_url, video_id, title, categories, safety_label)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (youtube_url)
                    DO UPDATE SET
                        video_id = EXCLUDED.video_id,
                        title = COALESCE(EXCLUDED.title, youtube_urls.title),
                        categories = COALESCE(EXCLUDED.categories, youtube_urls.categories),
                        safety_label = COALESCE(EXCLUDED.safety_label, youtube_urls.safety_label)
                    """,
                    (
                        db_url,
                        db_video_id,
                        db_title,
                        db_categories,
                        db_gt,
                    ),
                )
        conn.commit()
    finally:
        conn.close()

    return entries, get_pool_size()


def get_pool_size() -> int:
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM youtube_urls")
            return int(cur.fetchone()[0])
    finally:
        conn.close()


def get_pool_entries() -> list[PoolEntry]:
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, youtube_url, video_id, title, categories, safety_label
                FROM youtube_urls
                ORDER BY created_at ASC
                """
            )
            rows = cur.fetchall()
            return [
                PoolEntry(
                    id=str(r.get("id") or ""),
                    youtube_url=str(r.get("youtube_url") or ""),
                    video_id=str(r.get("video_id") or extract_video_id_from_url(str(r.get("youtube_url") or "") ) or ""),
                    title=(str(r.get("title")) if r.get("title") is not None else None),
                    categories=(str(r.get("categories")) if r.get("categories") is not None else None),
                    safety_label=_normalize_ground_truth_label(str(r.get("safety_label")) if r.get("safety_label") is not None else None),
                )
                for r in rows if r.get("youtube_url")
            ]
    finally:
        conn.close()


def coerce_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


def tokenize(text: str) -> set[str]:
    cleaned = re.sub(r"[^a-z0-9\s]", " ", (text or "").lower())
    words = [w for w in cleaned.split() if len(w) > 2]
    return set(words)


def average_ranks(values: list[float]) -> list[float]:
    indexed = sorted(enumerate(values), key=lambda x: x[1])
    ranks = [0.0] * len(values)
    i = 0
    while i < len(indexed):
        j = i
        while j + 1 < len(indexed) and indexed[j + 1][1] == indexed[i][1]:
            j += 1
        avg_rank = (i + j + 2) / 2.0
        for k in range(i, j + 1):
            ranks[indexed[k][0]] = avg_rank
        i = j + 1
    return ranks


def pearson_corr(xs: list[float], ys: list[float]) -> float | None:
    n = len(xs)
    if n < 2:
        return None
    mx = sum(xs) / n
    my = sum(ys) / n
    num = 0.0
    den_x = den_y = 0.0
    for x, y in zip(xs, ys):
        dx = x - mx
        dy = y - my
        num += dx * dy
        den_x += dx * dx
        den_y += dy * dy
    if den_x <= 0.0 or den_y <= 0.0:
        return None
    return num / math.sqrt(den_x * den_y)


def spearman_corr(xs: list[float], ys: list[float]) -> float:
    # For stable batch aggregation, invalid/no-variance cases return 0.0 instead of null.
    try:
        if len(xs) != len(ys) or len(xs) < 2:
            return 0.0
        corr = pearson_corr(average_ranks(xs), average_ranks(ys))
        return float(corr) if corr is not None else 0.0
    except Exception:
        return 0.0


def binary_metrics(tp: int, tn: int, fp: int, fn: int) -> tuple[float, float, float, float]:
    total = tp + tn + fp + fn
    accuracy = ((tp + tn) / total * 100.0) if total > 0 else 0.0
    precision = (tp / (tp + fp) * 100.0) if (tp + fp) > 0 else 0.0
    recall = (tp / (tp + fn) * 100.0) if (tp + fn) > 0 else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0.0
    return round(accuracy, 2), round(precision, 2), round(recall, 2), round(f1, 2)


def map_predicted_label(report: dict[str, Any]) -> str | None:
    verdict = str(report.get("verdict") or "").strip().upper()
    if verdict == "SAFE":
        return "SAFE"
    if verdict in {"UNSAFE", "PROFESSIONAL_REQUIRED"}:
        return "UNSAFE" if verdict == "UNSAFE" else "PSUA"

    risk = coerce_float(report.get("overall_risk_score"), default=float("nan"))
    if not math.isnan(risk):
        if risk >= 4.0:
            return "UNSAFE"
        if risk >= 3.0:
            return "PSUA"
        return "SAFE"
    return None


def evaluate_single_output(
    video_id: str,
    video_url: str,
    title: str,
    channel: str,
    scan_id: int | None,
    output_json: dict[str, Any],
    ground_truth_label: str | None,
) -> VideoEvalResult:
    try:
        from rule_extraction.extract_rules import override_severity
    except ModuleNotFoundError:
        sys.path.insert(0, str(ROOT))
        from rule_extraction.extract_rules import override_severity

    steps = output_json.get("steps") or []
    model_reports = output_json.get("modelReports") or {}
    qwen = model_reports.get("qwen") if isinstance(model_reports, dict) else None
    if not qwen:
        qwen = output_json.get("report") or {}

    step_analysis = qwen.get("step_safety_analysis") if isinstance(qwen, dict) else []
    if not isinstance(step_analysis, list):
        step_analysis = []

    by_step: dict[int, dict[str, Any]] = {}
    for s in step_analysis:
        if isinstance(s, dict) and isinstance(s.get("step_number"), int):
            by_step[int(s["step_number"])] = s

    tmp_rules = []
    for idx, step in enumerate(steps, start=1):
        if not isinstance(step, dict):
            continue
        step_num = step.get("step_number") if isinstance(step.get("step_number"), int) else idx
        txt = (step.get("step_text") or "").strip()
        action = (step.get("action_summary") or txt).strip()
        excerpt = (step.get("transcript_excerpt") or txt).strip()
        tmp_rules.append(
            {
                "rule_id": str(step_num),
                "suggested_severity": 1,
                "validated_severity": 1,
                "actionable_rule": action,
                "original_text": excerpt,
            }
        )

    scored_steps = override_severity(tmp_rules)
    override_by_step = {
        int(r.get("rule_id")): int(r.get("validated_severity") or 1)
        for r in scored_steps
        if str(r.get("rule_id", "")).isdigit()
    }

    tp = tn = fp = fn = 0
    scan_llm: list[float] = []
    scan_override: list[float] = []
    rr_list: list[float] = []
    total_precautions = 0
    supported_precautions = 0

    for idx, step in enumerate(steps, start=1):
        if not isinstance(step, dict):
            continue
        step_num = step.get("step_number") if isinstance(step.get("step_number"), int) else idx
        analysis = by_step.get(step_num, {})
        llm_score = coerce_float(analysis.get("risk_level"), default=float("nan"))
        if math.isnan(llm_score):
            continue
        override_score = float(override_by_step.get(step_num, 1))
        scan_llm.append(llm_score)
        scan_override.append(override_score)

        y_pred = 1 if llm_score >= 3.0 else 0
        y_true = 1 if override_score >= 3.0 else 0
        if y_true == 1 and y_pred == 1:
            tp += 1
        elif y_true == 0 and y_pred == 0:
            tn += 1
        elif y_true == 0 and y_pred == 1:
            fp += 1
        else:
            fn += 1

        matched_rules = analysis.get("matched_rules") or []
        rr = 0.0
        if isinstance(matched_rules, list) and matched_rules:
            for rank, rule in enumerate(matched_rules, start=1):
                if not isinstance(rule, dict):
                    continue
                sev = coerce_float(rule.get("severity"), default=0.0)
                if sev >= max(3.0, override_score):
                    rr = 1.0 / rank
                    break
        rr_list.append(rr)

        required_precautions = analysis.get("required_precautions") or []
        if isinstance(required_precautions, list):
            evidence_text = " ".join(
                [
                    str(step.get("step_text") or ""),
                    str(step.get("transcript_excerpt") or ""),
                    " ".join(
                        str((r.get("rule_text") or r.get("actionable_rule") or ""))
                        for r in matched_rules if isinstance(r, dict)
                    ),
                ]
            )
            evidence_tokens = tokenize(evidence_text)
            for precaution in required_precautions:
                tokens = tokenize(str(precaution))
                if not tokens:
                    continue
                total_precautions += 1
                overlap = len(tokens & evidence_tokens)
                if overlap >= 2 and (overlap / len(tokens)) >= 0.5:
                    supported_precautions += 1

    step_accuracy, step_precision, step_recall, step_f1_score = binary_metrics(tp, tn, fp, fn)
    mrr = round(sum(rr_list) / len(rr_list), 4) if rr_list else 0.0
    faithfulness = round((supported_precautions / total_precautions) * 100.0, 2) if total_precautions > 0 else 100.0
    spearman = spearman_corr(scan_llm, scan_override)
    avg_override_severity = round(sum(scan_override) / len(scan_override), 4) if scan_override else 0.0

    report = output_json.get("report") if isinstance(output_json.get("report"), dict) else {}
    predicted_label = map_predicted_label(report)
    risk_score = coerce_float(report.get("overall_risk_score"), default=0.0)

    gt_norm = _normalize_ground_truth_label(ground_truth_label)
    gt_bin = _ground_truth_to_binary(gt_norm)
    pred_bin = _ground_truth_to_binary(predicted_label)
    label_match = None
    label_tp = label_tn = label_fp = label_fn = 0
    if gt_bin is not None and pred_bin is not None:
        label_match = bool(gt_bin == pred_bin)
        if gt_bin == 1 and pred_bin == 1:
            label_tp = 1
        elif gt_bin == 0 and pred_bin == 0:
            label_tn = 1
        elif gt_bin == 0 and pred_bin == 1:
            label_fp = 1
        else:
            label_fn = 1

    return VideoEvalResult(
        url_id="",
        video_id=video_id,
        video_url=video_url,
        title=title,
        channel=channel,
        scan_id=scan_id,
        ground_truth_label=gt_norm,
        predicted_label=predicted_label,
        label_match=label_match,
        steps_evaluated=len(scan_llm),
        total_precautions=total_precautions,
        supported_precautions=supported_precautions,
        step_accuracy=step_accuracy,
        step_precision=step_precision,
        step_recall=step_recall,
        step_f1_score=step_f1_score,
        label_tp=label_tp,
        label_tn=label_tn,
        label_fp=label_fp,
        label_fn=label_fn,
        mrr=mrr,
        faithfulness=faithfulness,
        spearman=round(spearman, 4),
        risk_score=risk_score,
        average_severity_score=avg_override_severity,
    )


def _get_supabase_project_ref() -> str | None:
    url = (os.getenv("SUPABASE_URL") or "").strip()
    m = re.search(r"@db\.([^.]+)\.supabase\.co", url)
    if m:
        return m.group(1)
    m = re.search(r"postgres\.([^:]+):", url)
    if m:
        return m.group(1)
    return (os.getenv("SUPABASE_PROJECT_REF") or "").strip() or None


def _upload_batch_urls_json(payload: dict[str, Any]) -> str | None:
    try:
        project_ref = _get_supabase_project_ref()
        if not project_ref:
            return None
        api_key = (
            os.getenv("SUPABASE_SERVICE_KEY")
            or os.getenv("SUPABASE_ANON_KEY")
            or os.getenv("SUPABASE_KEY")
        )
        if not api_key:
            return None

        bucket = "system-eval"
        ts = int(time.time())
        object_name = f"batch_urls_{ts}.json"
        storage_url = f"https://{project_ref}.supabase.co/storage/v1/object/{bucket}/{object_name}"
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")

        with httpx.Client(timeout=30.0) as client:
            resp = client.post(
                storage_url,
                content=body,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "apikey": api_key,
                    "Content-Type": "application/json",
                    "x-upsert": "true",
                },
            )
            if resp.status_code not in (200, 201):
                return None

        return f"https://{project_ref}.supabase.co/storage/v1/object/public/{bucket}/{object_name}"
    except Exception:
        return None


def stream_analyze(api_base: str, video_id: str) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    extraction_obj: dict[str, Any] = {}
    reports: dict[str, dict[str, Any]] = {}
    comparison: dict[str, Any] | None = None

    with httpx.Client(timeout=300.0) as client:
        with client.stream("GET", f"{api_base}/api/analyze", params={"video_id": video_id}) as resp:
            resp.raise_for_status()
            for line in resp.iter_lines():
                if not line:
                    continue
                text = line.decode("utf-8", errors="ignore") if isinstance(line, bytes) else line
                if not text.startswith("data: "):
                    continue
                payload = text[6:]
                try:
                    event = json.loads(payload)
                except Exception:
                    continue
                e_type = event.get("type")
                if e_type == "error":
                    raise RuntimeError(str(event.get("message") or "Analyze route failed"))
                if e_type == "metadata":
                    metadata = event
                elif e_type == "steps_complete":
                    steps_json = event.get("steps_json") or "{}"
                    extraction_obj = json.loads(steps_json)
                elif e_type == "safety_report":
                    key = str(event.get("model_key") or "unknown")
                    report_json = event.get("report_json") or "{}"
                    reports[key] = json.loads(report_json)
                elif e_type == "model_comparison":
                    comp = event.get("comparison_json")
                    if isinstance(comp, str):
                        comparison = json.loads(comp)
                    elif isinstance(comp, dict):
                        comparison = comp

    if not extraction_obj:
        raise RuntimeError(f"No extraction result produced for video {video_id}")

    steps = extraction_obj.get("steps") or []
    safety_categories = extraction_obj.get("safety_categories") or []
    primary = reports.get("qwen") or (next(iter(reports.values())) if reports else {})
    if not primary:
        raise RuntimeError(f"No safety report produced for video {video_id}")

    return {
        "title": metadata.get("title") or extraction_obj.get("title") or f"Video {video_id}",
        "channel": metadata.get("author") or "",
        "output_json": {
            "extraction": extraction_obj,
            "steps": steps,
            "safetyCategories": safety_categories,
            "report": primary,
            "modelReports": reports,
            "comparison": comparison or {},
        },
    }


def _is_rate_limit_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return ("rate limit" in msg) or ("429" in msg) or ("too many requests" in msg)


def analyze_with_retries(api_base: str, video_id: str, max_attempts: int = 5) -> dict[str, Any]:
    last_exc: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            return stream_analyze(api_base, video_id)
        except Exception as exc:
            last_exc = exc
            if not _is_rate_limit_error(exc) or attempt == max_attempts:
                raise
            wait_seconds = min(60, 2 ** attempt)
            print(f"Rate limit hit for {video_id}; retrying in {wait_seconds}s (attempt {attempt}/{max_attempts})")
            time.sleep(wait_seconds)
    if last_exc:
        raise last_exc
    raise RuntimeError("Unknown analyze failure")


def assert_api_reachable(api_base: str) -> None:
    health_url = f"{api_base.rstrip('/')}/api/health"
    try:
        with httpx.Client(timeout=8.0) as client:
            resp = client.get(health_url)
            if resp.status_code != 200:
                raise RuntimeError(f"health check returned HTTP {resp.status_code}")
    except Exception as exc:
        raise RuntimeError(
            "Backend API is not reachable at "
            f"{health_url}. Start the backend first (example: `python backend/app.py`) "
            "or pass the correct --api-base URL. "
            f"Original error: {exc}"
        )


def _is_api_up(api_base: str) -> bool:
    health_url = f"{api_base.rstrip('/')}/api/health"
    try:
        with httpx.Client(timeout=3.0) as client:
            resp = client.get(health_url)
            return resp.status_code == 200
    except Exception:
        return False


def _looks_local_api(api_base: str) -> bool:
    base = api_base.lower()
    return (
        "127.0.0.1" in base
        or "localhost" in base
        or "0.0.0.0" in base
    )


def ensure_api_ready(api_base: str) -> None:
    if _is_api_up(api_base):
        return

    if not _looks_local_api(api_base):
        assert_api_reachable(api_base)
        return

    backend_app = ROOT / "backend" / "app.py"
    if not backend_app.exists():
        assert_api_reachable(api_base)
        return

    print("Backend API is not running. Auto-starting backend...")
    proc = subprocess.Popen(
        [sys.executable, str(backend_app)],
        cwd=str(ROOT / "backend"),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    def _cleanup() -> None:
        try:
            if proc.poll() is None:
                proc.terminate()
        except Exception:
            pass

    atexit.register(_cleanup)

    deadline = time.time() + 30
    while time.time() < deadline:
        if _is_api_up(api_base):
            print("Backend API started successfully.")
            return
        if proc.poll() is not None:
            break
        time.sleep(1)

    assert_api_reachable(api_base)


def save_scan(api_base: str, video_id: str, video_url: str, title: str, channel: str, output_json: dict[str, Any]) -> int | None:
    report = output_json.get("report") or {}
    body = {
        "video_id": video_id,
        "video_url": video_url,
        "title": title,
        "channel": channel,
        "verdict": report.get("verdict") or "",
        "risk_score": report.get("overall_risk_score"),
        "output_json": output_json,
    }
    try:
        with httpx.Client(timeout=60.0) as client:
            resp = client.post(f"{api_base}/api/scans", json=body)
            resp.raise_for_status()
            return int(resp.json().get("id"))
    except Exception:
        return None


def persist_batch_result(
    selected_entries: list[PoolEntry],
    per_video: list[VideoEvalResult],
    from_index: int,
    to_index: int,
) -> dict[str, Any]:
    if not per_video:
        raise RuntimeError("No video metrics to persist")

    tp = sum(v.label_tp for v in per_video)
    tn = sum(v.label_tn for v in per_video)
    fp = sum(v.label_fp for v in per_video)
    fn = sum(v.label_fn for v in per_video)
    labeled_count = tp + tn + fp + fn

    label_accuracy, label_precision, label_recall, label_f1 = binary_metrics(tp, tn, fp, fn)

    avg_mrr = round(sum(v.mrr for v in per_video) / len(per_video), 4)
    avg_faith = round(sum(v.faithfulness for v in per_video) / len(per_video), 2)
    avg_spearman = round(sum(v.spearman for v in per_video) / len(per_video), 4) if per_video else 0.0
    selected_urls = [e.youtube_url for e in selected_entries]
    bucket_source_url = _upload_batch_urls_json(
        {
            "from_index": from_index,
            "to_index": to_index,
            "urls": selected_urls,
        }
    )

    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO system_eval
                    (batch_start_index, batch_end_index, bucket_source_url,
                     accuracy, precision, recall, f1_score,
                     avg_mrr, avg_faithfulness, avg_spearman)
                VALUES (%s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s)
                RETURNING id, evaluated_at
                """,
                (
                    from_index,
                    to_index,
                    bucket_source_url,
                    label_accuracy,
                    label_precision,
                    label_recall,
                    label_f1,
                    avg_mrr,
                    avg_faith,
                    avg_spearman,
                ),
            )
            inserted = cur.fetchone()
            eval_id = str(inserted["id"])

            for v in per_video:
                cur.execute(
                    """
                    INSERT INTO individual_video_results
                        (eval_id, url_id, predicted_label, risk_score,
                         total_steps, average_severity_score,
                         spearman_correlation, faithfulness_score)
                    VALUES (%s, %s, %s, %s,
                            %s, %s,
                            %s, %s)
                    """,
                    (
                        eval_id,
                        v.url_id,
                        v.predicted_label,
                        v.risk_score,
                        v.steps_evaluated,
                        v.average_severity_score,
                        v.spearman,
                        v.faithfulness,
                    ),
                )

        conn.commit()
        return {
            "eval_id": eval_id,
            "evaluated_at": inserted["evaluated_at"].isoformat() if inserted and inserted.get("evaluated_at") else datetime.now(timezone.utc).isoformat(),
            "evaluated_scans": len(per_video),
            "selected_urls_count": len(selected_entries),
            "batch_start_index": from_index,
            "batch_end_index": to_index,
            "bucket_source_url": bucket_source_url,
            "metrics": {
                "accuracy": label_accuracy,
                "precision": label_precision,
                "recall": label_recall,
                "f1_score": label_f1,
                "mean_reciprocal_rank": avg_mrr,
                "faithfulness_score": avg_faith,
                "spearman_correlation": avg_spearman,
            },
            "labeled_videos_in_batch": labeled_count,
        }
    finally:
        conn.close()


def run_batch(
    api_base: str,
    urls_dir: Path,
    from_index: int,
    to_index: int,
    pool: list[PoolEntry] | None = None,
    added_entries: list[PoolEntry] | None = None,
    total_pool_after: int | None = None,
) -> dict[str, Any]:
    ensure_api_ready(api_base)
    ensure_system_eval_schema()
    if pool is None:
        added_entries, total_pool_after = collect_and_upsert_urls(urls_dir)
        pool = get_pool_entries()
    if added_entries is None:
        added_entries = []
    if total_pool_after is None:
        total_pool_after = len(pool)

    if not pool:
        raise RuntimeError("youtube_urls pool is empty. Add files into system_eval/youtube_urls first.")

    if from_index < 1:
        raise RuntimeError("--from-index must be >= 1")
    if to_index < from_index:
        raise RuntimeError("--to-index must be >= --from-index")
    if to_index > len(pool):
        raise RuntimeError(f"--to-index ({to_index}) exceeds total URLs in pool ({len(pool)})")

    selected_entries = pool[from_index - 1:to_index]
    if len(selected_entries) < 5:
        raise RuntimeError("Batch size must be at least 5 videos (selected range is smaller).")

    labeled_in_selection = sum(1 for e in selected_entries if e.safety_label)
    if labeled_in_selection == 0:
        raise RuntimeError(
            "Selected batch has 0 labeled rows in youtube_urls.safety_label. "
            "Classification metrics (accuracy/precision/recall/f1) cannot be computed. "
            "Fix label ingestion first (expected labels: SAFE/UNSAFE/PSUA)."
        )

    selected_urls = [e.youtube_url for e in selected_entries]

    print(f"Total URLs in pool: {len(pool)}")
    print(f"New/updated URLs from folder: {len(added_entries)}")
    print(f"Selected range: {from_index}..{to_index} ({len(selected_urls)} urls)")
    print(f"Labeled rows in selected range: {labeled_in_selection}")

    per_video: list[VideoEvalResult] = []
    failures: list[dict[str, str]] = []

    for idx, entry in enumerate(selected_entries, start=from_index):
        url = entry.youtube_url
        video_id = entry.video_id or extract_video_id_from_url(url)
        if not video_id:
            failures.append({"url": url, "error": "Invalid YouTube URL"})
            continue

        print(f"[{idx}] Analyzing {video_id} ...")
        try:
            result = analyze_with_retries(api_base, video_id)
            output_json = result["output_json"]
            title = str(result.get("title") or f"Video {video_id}")
            channel = str(result.get("channel") or "")
            scan_id = save_scan(api_base, video_id, url, title, channel, output_json)
            evaluated = evaluate_single_output(
                    video_id,
                    url,
                    title,
                    channel,
                    scan_id,
                    output_json,
                    entry.safety_label,
                )
            evaluated.url_id = entry.id
            per_video.append(evaluated)
            print(f"[{idx}] done | title={title}")
        except Exception as exc:
            failures.append({"url": url, "error": str(exc)})
            print(f"[{idx}] failed | {exc}")

    if not per_video:
        raise RuntimeError(f"No videos evaluated successfully. Failures: {len(failures)}")

    summary = persist_batch_result(selected_entries, per_video, from_index, to_index)
    summary["failures"] = failures
    summary["per_video"] = [
        {
            "url_id": v.url_id,
            "video_id": v.video_id,
            "video_url": v.video_url,
            "title": v.title,
            "scan_id": v.scan_id,
            "ground_truth_label": v.ground_truth_label,
            "predicted_label": v.predicted_label,
            "label_match": v.label_match,
            "metrics": {
                "risk_score": v.risk_score,
                "average_severity_score": v.average_severity_score,
                "mean_reciprocal_rank": v.mrr,
                "faithfulness_score": v.faithfulness,
                "spearman_correlation": v.spearman,
            },
            "total_steps": v.steps_evaluated,
        }
        for v in per_video
    ]
    summary["total_urls_in_pool"] = total_pool_after
    summary["batch_size"] = len(selected_entries)
    return summary


def _prompt_index(prompt: str, min_value: int, max_value: int) -> int:
    while True:
        raw = input(prompt).strip()
        try:
            value = int(raw)
        except Exception:
            print("Please enter a valid integer.")
            continue
        if value < min_value or value > max_value:
            print(f"Value must be between {min_value} and {max_value}.")
            continue
        return value


def main() -> int:
    parser = argparse.ArgumentParser(description="Batch system evaluation from URL pool")
    parser.add_argument("--api-base", default="http://127.0.0.1:8000", help="Backend base URL")
    parser.add_argument("--urls-dir", default=str(Path(__file__).resolve().parent / "youtube_urls"), help="Folder containing PDF/CSV/XLSX/TXT with URLs")
    parser.add_argument("--from-index", type=int, help="1-based start index in youtube_urls pool")
    parser.add_argument("--to-index", type=int, help="1-based end index in youtube_urls pool")
    parser.add_argument("--collect-only", action="store_true", help="Only collect URLs into youtube_urls table, do not run analysis")
    parser.add_argument("--run-only", action="store_true", help="Run evaluation using existing DB URL pool without collecting files first")
    parser.add_argument("--verify-db", action="store_true", help="Print verification summary from Supabase tables")
    parser.add_argument("--out", default=str(Path(__file__).resolve().parent / "last_batch_result.json"), help="Output result JSON path")
    args = parser.parse_args()

    if args.collect_only and args.run_only:
        raise SystemExit("--collect-only and --run-only cannot be used together")

    if args.collect_only:
        urls_dir = Path(args.urls_dir)
        added_entries, total_pool_after = collect_and_upsert_urls(urls_dir)
        summary = {
            "collect_only": True,
            "added_or_updated_urls": len(added_entries),
            "added_or_updated_labeled": sum(1 for e in added_entries if e.safety_label),
            "total_urls_in_pool": total_pool_after,
            "sample_urls": [e.youtube_url for e in added_entries[:10]],
        }
        if args.verify_db:
            summary["verification"] = verify_supabase_state()
        print(json.dumps(summary, indent=2))
        return 0

    urls_dir = Path(args.urls_dir)
    if args.run_only:
        ensure_system_eval_schema()
        added_entries = []
        pool = get_pool_entries()
        total_pool_after = len(pool)
    else:
        added_entries, total_pool_after = collect_and_upsert_urls(urls_dir)
        pool = get_pool_entries()

    if not pool:
        raise SystemExit("youtube_urls pool is empty after collection. Add files and try again.")

    if args.run_only:
        print("Run-only mode: skipped file collection.")
    else:
        print(f"Collected/updated URLs: {len(added_entries)}")
    print(f"Total URLs in pool: {len(pool)}")

    from_index = args.from_index
    to_index = args.to_index

    if from_index is None:
        from_index = _prompt_index(
            f"Enter start index (1 to {len(pool)}): ",
            1,
            len(pool),
        )
    if to_index is None:
        to_index = _prompt_index(
            f"Enter end index ({from_index} to {len(pool)}): ",
            from_index,
            len(pool),
        )

    summary = run_batch(
        api_base=args.api_base.rstrip("/"),
        urls_dir=urls_dir,
        from_index=from_index,
        to_index=to_index,
        pool=pool,
        added_entries=added_entries,
        total_pool_after=total_pool_after,
    )

    if args.verify_db:
        summary["verification"] = verify_supabase_state()

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print("\nBatch evaluation complete")
    print(json.dumps({
        "eval_id": summary.get("eval_id"),
        "evaluated_scans": summary.get("evaluated_scans"),
        "metrics": summary.get("metrics"),
        "failures": len(summary.get("failures", [])),
    }, indent=2))
    print(f"Saved: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
